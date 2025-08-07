#!/usr/bin/env python3
"""
Debug script to analyze Excel file and identify missing text extraction
"""

import logging
from offitrans.processors.excel import ExcelProcessor
from offitrans.core.config import Config
from offitrans.translators.google import GoogleTranslator
from openpyxl import load_workbook

# Setup logging
logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

def analyze_excel_file(file_path: str):
    """Analyze Excel file for comprehensive text extraction"""
    print(f"=== Analyzing Excel file: {file_path} ===\n")
    
    try:
        # Method 1: Use our enhanced processor
        print("1. Using Enhanced Offitrans Processor:")
        print("-" * 50)
        
        config = Config()
        translator = GoogleTranslator()
        processor = ExcelProcessor(translator=translator, config=config)
        
        text_data = processor.extract_text(file_path)
        
        print(f"Found {len(text_data)} text entries using Offitrans processor:")
        for i, item in enumerate(text_data):
            print(f"{i+1:3d}. {item['sheet_name']}!{item['cell_coordinate']}: '{item['text'][:50]}...'")
        
        print(f"\n--- End of Offitrans extraction ({len(text_data)} items) ---\n")
        
        # Method 2: Direct openpyxl analysis for comparison
        print("2. Direct openpyxl Analysis (All Cells):")
        print("-" * 50)
        
        workbook = load_workbook(file_path, data_only=False)
        all_texts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            
            cell_count = 0
            text_count = 0
            
            # Check ALL cells in the used range
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_count += 1
                    
                    # Check if cell has any value
                    if cell.value is not None:
                        cell_value = str(cell.value).strip()
                        if cell_value:  # Non-empty string
                            text_count += 1
                            coordinate = cell.coordinate
                            
                            # Check if this text was found by our processor
                            found_in_processor = any(
                                item['cell_coordinate'] == coordinate and item['text'].strip() == cell_value
                                for item in text_data
                            )
                            
                            status = "✓ Found" if found_in_processor else "✗ MISSING"
                            
                            print(f"  {coordinate}: '{cell_value[:50]}...' [{status}]")
                            
                            if not found_in_processor:
                                print(f"    → Cell type: {type(cell.value)}")
                                print(f"    → Cell._value type: {type(cell._value) if hasattr(cell, '_value') else 'None'}")
                                print(f"    → Starts with '=': {cell_value.startswith('=')}")
                                print(f"    → Is string: {isinstance(cell.value, str)}")
                                print(f"    → After strip: '{cell_value}' (len={len(cell_value)})")
                                
                                # Check if it's a formula result
                                if hasattr(cell, 'data_type'):
                                    print(f"    → Data type: {cell.data_type}")
                            
                            all_texts.append({
                                'sheet': sheet_name,
                                'coordinate': coordinate,
                                'text': cell_value,
                                'found_by_processor': found_in_processor,
                                'cell_type': type(cell.value).__name__,
                                'is_formula': cell_value.startswith('=') if isinstance(cell.value, str) else False
                            })
            
            print(f"  Total cells checked: {cell_count}")
            print(f"  Cells with text: {text_count}")
        
        workbook.close()
        
        # Summary
        print("\n3. Summary:")
        print("-" * 50)
        total_texts = len(all_texts)
        found_by_processor = sum(1 for item in all_texts if item['found_by_processor'])
        missing_texts = [item for item in all_texts if not item['found_by_processor']]
        
        print(f"Total text cells found by direct analysis: {total_texts}")
        print(f"Text cells found by Offitrans processor: {found_by_processor}")
        print(f"Missing text cells: {len(missing_texts)}")
        
        if missing_texts:
            print(f"\nMissing texts analysis:")
            formulas = [item for item in missing_texts if item['is_formula']]
            non_formulas = [item for item in missing_texts if not item['is_formula']]
            
            print(f"  - Formula cells: {len(formulas)}")
            print(f"  - Non-formula cells: {len(non_formulas)}")
            
            if non_formulas:
                print(f"\nNon-formula missing texts (first 10):")
                for item in non_formulas[:10]:
                    print(f"  {item['sheet']}!{item['coordinate']}: '{item['text'][:30]}...' (type: {item['cell_type']})")
                    
        print(f"\nExtraction efficiency: {found_by_processor/total_texts*100:.1f}%")
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    sample_file = "/root/projects/github/Offitrans/examples/sample_files/sample.xlsx"
    analyze_excel_file(sample_file)