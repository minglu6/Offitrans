"""
Excel file processor for Offitrans

This module provides functionality to translate Excel files while preserving
formatting, images, and layout.
"""

import os
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.styles.colors import Color
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image as PILImage

    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

from .base import BaseProcessor
from ..exceptions.errors import ExcelProcessorError

logger = logging.getLogger(__name__)


class ExcelProcessor(BaseProcessor):
    """
    Excel file processor that handles translation while preserving formatting.

    This processor can handle complex Excel files with:
    - Rich text formatting
    - Merged cells
    - Images and charts
    - Multiple worksheets
    - Various cell formats and styles
    """

    def __init__(self, **kwargs):
        """
        Initialize Excel processor.

        Args:
            **kwargs: Additional arguments passed to BaseProcessor
        """
        if not OPENPYXL_AVAILABLE:
            raise ExcelProcessorError(
                "openpyxl library is required for Excel processing",
                details="Install with: pip install openpyxl",
            )

        super().__init__(**kwargs)

        # Excel-specific settings
        self.smart_column_width = getattr(
            self.config.processor, "smart_column_width", True
        )

        # Image data storage
        self.image_data: Dict[str, List[Dict[str, Any]]] = {}

    def supports_file_type(self, file_path: str) -> bool:
        """
        Check if file type is supported.

        Args:
            file_path: Path to the file

        Returns:
            True if file type is supported
        """
        supported_extensions = {".xlsx", ".xls", ".xlsm"}
        return Path(file_path).suffix.lower() in supported_extensions

    def extract_images_info(self, workbook) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract image information from Excel workbook.

        Args:
            workbook: openpyxl workbook object

        Returns:
            Dictionary mapping sheet names to image info lists
        """
        images_info = {}

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_images = []

                # Check for images in the sheet
                if hasattr(sheet, "_images") and sheet._images:
                    logger.info(
                        f"Found {len(sheet._images)} images in sheet '{sheet_name}'"
                    )

                    for img in sheet._images:
                        img_info = {
                            "image_object": img,
                            "anchor_type": type(img.anchor).__name__,
                        }

                        # Extract anchor information
                        if isinstance(img.anchor, TwoCellAnchor):
                            img_info["anchor_info"] = {
                                "type": "two_cell",
                                "from_col": img.anchor._from.col,
                                "from_col_off": img.anchor._from.colOff,
                                "from_row": img.anchor._from.row,
                                "from_row_off": img.anchor._from.rowOff,
                                "to_col": img.anchor.to.col,
                                "to_col_off": img.anchor.to.colOff,
                                "to_row": img.anchor.to.row,
                                "to_row_off": img.anchor.to.rowOff,
                            }
                        elif isinstance(img.anchor, OneCellAnchor):
                            img_info["anchor_info"] = {
                                "type": "one_cell",
                                "from_col": img.anchor._from.col,
                                "from_col_off": img.anchor._from.colOff,
                                "from_row": img.anchor._from.row,
                                "from_row_off": img.anchor._from.rowOff,
                                "width": img.anchor.ext.cx,
                                "height": img.anchor.ext.cy,
                            }

                        sheet_images.append(img_info)

                images_info[sheet_name] = sheet_images

        except Exception as e:
            logger.error(f"Error extracting image info: {e}")

        return images_info

    def restore_images_info(
        self, workbook, images_info: Dict[str, List[Dict[str, Any]]]
    ) -> None:
        """
        Restore image information to Excel workbook with enhanced error handling.

        Args:
            workbook: openpyxl workbook object
            images_info: Dictionary with image information
        """
        try:
            for sheet_name, sheet_images in images_info.items():
                if not sheet_images:
                    continue

                sheet = workbook[sheet_name]

                # Clear existing images
                if hasattr(sheet, "_images"):
                    sheet._images.clear()
                else:
                    sheet._images = []

                # Restore images
                for img_info in sheet_images:
                    try:
                        img_obj = img_info["image_object"]

                        # Use safe image creation method
                        new_img = self._safe_create_image(img_obj)
                        if new_img is None:
                            logger.warning("Could not create image object, skipping this image")
                            continue

                        # Restore anchor information
                        anchor_info = img_info.get("anchor_info", {})
                        if anchor_info.get("type") == "two_cell":
                            # Create TwoCellAnchor
                            anchor = TwoCellAnchor()
                            anchor._from.col = anchor_info["from_col"]
                            anchor._from.colOff = anchor_info["from_col_off"]
                            anchor._from.row = anchor_info["from_row"]
                            anchor._from.rowOff = anchor_info["from_row_off"]
                            anchor.to.col = anchor_info["to_col"]
                            anchor.to.colOff = anchor_info["to_col_off"]
                            anchor.to.row = anchor_info["to_row"]
                            anchor.to.rowOff = anchor_info["to_row_off"]

                        elif anchor_info.get("type") == "one_cell":
                            # Create OneCellAnchor
                            anchor = OneCellAnchor()
                            anchor._from.col = anchor_info["from_col"]
                            anchor._from.colOff = anchor_info["from_col_off"]
                            anchor._from.row = anchor_info["from_row"]
                            anchor._from.rowOff = anchor_info["from_row_off"]
                            anchor.ext.cx = anchor_info["width"]
                            anchor.ext.cy = anchor_info["height"]
                        else:
                            # Use original anchor
                            anchor = img_obj.anchor

                        new_img.anchor = anchor
                        try:
                            sheet.add_image(new_img)
                            logger.debug(f"Successfully added image to sheet {sheet_name}")
                        except Exception as add_err:
                            logger.warning(f"Adding image to sheet failed: {add_err}")
                            # Try using default anchor to re-add
                            try:
                                default_anchor = OneCellAnchor()
                                new_img.anchor = default_anchor
                                sheet.add_image(new_img)
                                logger.debug("Successfully added image using default anchor")
                            except Exception as default_err:
                                logger.error(f"Using default anchor also failed: {default_err}")
                                continue

                    except Exception as e:
                        logger.error(f"Error restoring image: {e}")
                        # If unable to restore anchor, try alternative approach
                        try:
                            logger.debug("Trying to use original image object...")
                            # Check original image object status
                            if hasattr(img_obj, "anchor") and img_obj.anchor:
                                sheet.add_image(img_obj)
                                logger.debug("Successfully used original image object")
                            else:
                                # Create a simple default anchor
                                default_anchor = OneCellAnchor()
                                default_anchor._from.col = 0
                                default_anchor._from.row = 0
                                default_anchor._from.colOff = 0
                                default_anchor._from.rowOff = 0

                                # Set default size
                                default_anchor.ext.cx = 2000000  # Default width
                                default_anchor.ext.cy = 2000000  # Default height

                                img_obj.anchor = default_anchor
                                sheet.add_image(img_obj)
                                logger.debug("Successfully used default anchor")
                        except Exception as fallback_err:
                            logger.error(f"All image restoration methods failed: {fallback_err}")
                            logger.info("Skipping this image, continuing with others")
                            continue

        except Exception as e:
            logger.error(f"Error restoring images: {e}")
    
    def _safe_create_image(self, img_obj) -> Optional[Image]:
        """
        Safely create image object, handling various possible errors.
        
        Args:
            img_obj: Original image object
            
        Returns:
            New image object or None
        """
        try:
            # Method 1: Direct use of original object (safest)
            if hasattr(img_obj, "anchor"):
                logger.debug("Using original image object (recommended method)")
                return img_obj
            
            # Method 2: Try using _data() method
            if hasattr(img_obj, "_data"):
                try:
                    img_data = img_obj._data()
                    if img_data:
                        # Check and clean data
                        if isinstance(img_data, bytes):
                            # Remove null bytes
                            if b'\x00' in img_data:
                                logger.debug("Detected null bytes, cleaning...")
                                img_data = img_data.replace(b'\x00', b'')
                            
                            # Validate image data (if PIL available)
                            if PIL_AVAILABLE:
                                try:
                                    # Use PIL to validate image data
                                    import io
                                    test_img = PILImage.open(io.BytesIO(img_data))
                                    test_img.verify()
                                    logger.debug("Image data validation successful")
                                except Exception as pil_err:
                                    logger.debug(f"PIL validation failed: {pil_err}")
                                    # Continue trying to use data
                            else:
                                logger.debug("Skipping PIL validation (not installed)")
                            
                            # Create new openpyxl image object
                            try:
                                new_img = Image(img_data)
                                logger.debug("Successfully created image using cleaned data")
                                return new_img
                            except Exception as create_err:
                                logger.debug(f"Failed to create image using cleaned data: {create_err}")
                                pass
                        
                except Exception as data_err:
                    logger.debug(f"Failed to get image data: {data_err}")
            
            # Method 3: Try using other attributes
            if hasattr(img_obj, "ref"):
                try:
                    logger.debug("Trying to use image reference")
                    # This may need to reload image from workbook
                    return img_obj
                except Exception:
                    pass
            
            # If all methods fail, return original object
            logger.debug("All methods failed, returning original object")
            return img_obj
            
        except Exception as e:
            logger.error(f"Image object creation completely failed: {e}")
            return None

    def extract_text(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract text content from Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of dictionaries containing text and metadata
        """
        text_data = []

        try:
            workbook = load_workbook(file_path, data_only=False)
            logger.info(f"Successfully opened Excel file: {file_path}")

            # Extract image information if image protection is enabled
            if self.image_protection:
                logger.info("Extracting image information...")
                self.image_data = self.extract_images_info(workbook)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"Processing worksheet: {sheet_name}")

                # Iterate through all cells
                for row in sheet.iter_rows():
                    for cell in row:
                        if (
                            cell.value
                            and isinstance(cell.value, str)
                            and cell.value.strip()
                        ):
                            # Skip formula cells (starting with =)
                            if not cell.value.startswith("="):
                                # Extract cell format information
                                format_info = self._extract_cell_format(cell)

                                # Check for rich text formatting
                                rich_text_info = self._extract_rich_text_format(cell)

                                text_data.append(
                                    {
                                        "text": cell.value,
                                        "sheet_name": sheet_name,
                                        "row": cell.row,
                                        "column": cell.column,
                                        "cell_coordinate": cell.coordinate,
                                        "format_info": format_info,
                                        "rich_text_info": rich_text_info,
                                    }
                                )

                                logger.debug(
                                    f"Extracted text from {sheet_name}!{cell.coordinate}: '{cell.value[:50]}...'"
                                )
                                
                                # Special attention to row 78 columns M-Q (referenced in original code)
                                if cell.row == 78 and cell.column >= 13 and cell.column <= 17:  # M=13, Q=17
                                    logger.info(f"Special attention: Row 78 M-Q column {cell.coordinate}")
                                    logger.info(f"  Text content: '{cell.value}'")
                                    logger.info(f"  Rich text info: {rich_text_info}")
                                    
                                    # Detailed check of this cell
                                    logger.info(f"  Raw content check:")
                                    logger.info(f"    cell.value: {type(cell.value)} = {cell.value}")
                                    logger.info(f"    cell._value: {type(cell._value) if hasattr(cell, '_value') else 'None'}")
                                    
                                    # Check merged cell
                                    merged_info = self._check_merged_cell(cell)
                                    if merged_info:
                                        logger.info(f"  Merged cell info: {merged_info}")

            workbook.close()
            logger.info(f"Total extracted {len(text_data)} text cells")
            return text_data

        except Exception as e:
            raise ExcelProcessorError(
                f"Failed to extract text from Excel file",
                details=str(e),
                file_path=file_path,
            ) from e

    def translate_and_save(
        self, file_path: str, output_path: str, target_language: str = "en"
    ) -> bool:
        """
        Translate Excel file and save to output path.

        Args:
            file_path: Path to input Excel file
            output_path: Path for output Excel file
            target_language: Target language code

        Returns:
            True if successful, False otherwise
        """
        try:
            # Step 1: Extract text and metadata
            logger.info("Step 1: Extracting text from Excel file...")
            text_data = self.extract_text(file_path)

            if not text_data:
                logger.warning("No translatable text found in Excel file")
                return False

            # Step 2: Preprocess and translate texts
            logger.info("Step 2: Translating texts...")
            original_texts = [item["text"] for item in text_data]
            unique_texts, metadata = self.preprocess_texts(original_texts)
            translated_unique = self.translate_texts(unique_texts, target_language)
            translated_texts = self.postprocess_translations(
                original_texts, translated_unique, metadata
            )

            # Step 3: Apply translations to Excel file
            logger.info("Step 3: Applying translations to Excel file...")
            success = self._replace_text_with_format_and_images(
                file_path, output_path, text_data, translated_texts, target_language
            )

            if success:
                logger.info(f"Successfully translated Excel file: {output_path}")
                return True
            else:
                logger.error("Failed to apply translations to Excel file")
                return False

        except Exception as e:
            logger.error(f"Error translating Excel file: {e}")
            return False

    def _replace_text_with_format_and_images(
        self,
        excel_path: str,
        output_path: str,
        text_data: List[Dict[str, Any]],
        translated_texts: List[str],
        target_language: str = "en",
    ) -> bool:
        """
        Replace text in Excel file while preserving formatting and images.

        Args:
            excel_path: Input Excel file path
            output_path: Output Excel file path
            text_data: Original text data with metadata
            translated_texts: List of translated texts
            target_language: Target language code

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = load_workbook(excel_path, data_only=False)

            # Replace text in cells
            for item, translated_text in zip(text_data, translated_texts):
                sheet_name = item["sheet_name"]
                row = item["row"]
                column = item["column"]
                format_info = item["format_info"]

                # Get worksheet and cell
                sheet = workbook[sheet_name]
                cell = sheet.cell(row=row, column=column)

                # Replace text
                cell.value = translated_text

                # Apply formatting
                self._apply_cell_format(cell, format_info, target_language)

                # Apply rich text formatting if available
                rich_text_info = item.get("rich_text_info")
                if rich_text_info and rich_text_info.get("has_rich_text"):
                    self._apply_rich_text_format(
                        cell,
                        item["text"],
                        translated_text,
                        rich_text_info,
                        target_language,
                    )

                # Handle merged cell synchronization
                merged_cell_info = self._check_merged_cell(cell)
                if merged_cell_info:
                    logger.debug(f"Processing merged cell: {merged_cell_info['range']}")
                    self._synchronize_merged_cell_formats(cell, item["text"], translated_text, format_info, rich_text_info, merged_cell_info)
                
                # Special processing for row 78 M-Q columns (compatibility with reference code)
                if cell.row == 78 and cell.column >= 13 and cell.column <= 17:  # M=13, Q=17
                    logger.info(f"Special attention row 78 {cell.coordinate}")
                    logger.info(f"  Translation before: '{item['text']}'")
                    logger.info(f"  Translation after: '{translated_text}'")
                    logger.info(f"  Rich text info: {rich_text_info}")
                    
                    # If no rich text detected but may exist, try forced recheck
                    if not rich_text_info:
                        logger.info(f"  Forced rich text recheck...")
                        rich_text_info = self._extract_rich_text_format(cell)
                        if rich_text_info:
                            logger.info(f"  Recheck found rich text: {rich_text_info}")
                            self._apply_rich_text_format(
                                cell, item["text"], translated_text, rich_text_info, target_language
                            )
                            
                            # If found rich text and is merged cell, re-synchronize
                            if merged_cell_info:
                                self._synchronize_merged_cell_formats(cell, item["text"], translated_text, format_info, rich_text_info, merged_cell_info)
                
                logger.debug(f"Applied translation to {sheet_name}!{cell.coordinate}")

            # Restore images if image protection is enabled
            if self.image_protection and self.image_data:
                logger.info("Restoring image information...")
                self.restore_images_info(workbook, self.image_data)

            # Apply smart column width adjustment if enabled
            if self.smart_column_width:
                logger.info("Adjusting column widths...")
                self._smart_adjust_column_width(workbook)

            # Save the workbook
            workbook.save(output_path)
            workbook.close()

            logger.info(f"Successfully saved translated Excel file: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error replacing text in Excel file: {e}")
            return False

    def _extract_cell_format(self, cell) -> Dict[str, Any]:
        """
        Extract formatting information from a cell.

        Args:
            cell: openpyxl cell object

        Returns:
            Dictionary containing format information
        """
        format_info = {}

        try:
            # Font information
            if cell.font:
                format_info["font_name"] = cell.font.name
                format_info["font_size"] = cell.font.size
                format_info["font_bold"] = cell.font.bold
                format_info["font_italic"] = cell.font.italic
                format_info["font_underline"] = cell.font.underline
                format_info["font_strike"] = cell.font.strike

                # Color information
                if cell.font.color:
                    format_info["font_color"] = cell.font.color
                    if hasattr(cell.font.color, "rgb") and cell.font.color.rgb:
                        format_info["font_color_rgb"] = cell.font.color.rgb
                    elif (
                        hasattr(cell.font.color, "indexed")
                        and cell.font.color.indexed is not None
                    ):
                        format_info["font_color_indexed"] = cell.font.color.indexed
                    elif (
                        hasattr(cell.font.color, "theme")
                        and cell.font.color.theme is not None
                    ):
                        format_info["font_color_theme"] = cell.font.color.theme
                        if (
                            hasattr(cell.font.color, "tint")
                            and cell.font.color.tint is not None
                        ):
                            format_info["font_color_tint"] = cell.font.color.tint

            # Fill information
            if cell.fill and hasattr(cell.fill, "start_color"):
                format_info["fill_color"] = cell.fill.start_color
                format_info["fill_type"] = cell.fill.fill_type
                format_info["fill_object"] = cell.fill

            # Alignment information
            if cell.alignment:
                format_info["horizontal"] = cell.alignment.horizontal
                format_info["vertical"] = cell.alignment.vertical
                format_info["wrap_text"] = cell.alignment.wrap_text
                format_info["shrink_to_fit"] = cell.alignment.shrink_to_fit

            # Border information
            if cell.border:
                format_info["has_border"] = True
                format_info["border"] = cell.border

            # Number format
            if cell.number_format:
                format_info["number_format"] = cell.number_format

        except Exception as e:
            logger.error(f"Error extracting cell format: {e}")

        return format_info

    def _extract_rich_text_format(self, cell) -> Optional[Dict[str, Any]]:
        """
        Extract rich text formatting information from a cell.

        Args:
            cell: openpyxl cell object

        Returns:
            Rich text format information or None
        """
        try:
            # Enhanced debugging information
            cell_text = str(cell.value) if cell.value else ""
            logger.debug(f"Checking cell {cell.coordinate}: '{cell_text[:30]}...'")
            logger.debug(f"Cell type: {type(cell.value)}")
            logger.debug(f"_value type: {type(cell._value) if hasattr(cell, '_value') else 'None'}")
            
            # Check merged cell status
            merged_info = None
            if hasattr(cell, 'coordinate'):
                worksheet = cell.parent
                if worksheet and hasattr(worksheet, 'merged_cells'):
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            logger.debug(f"Detected merged cell: {merged_range}")
                            merged_info = {
                                'range': str(merged_range),
                                'top_left': merged_range.coord.split(':')[0]
                            }
                            break
            
            # Method 1: Check _value attribute
            if hasattr(cell, '_value') and isinstance(cell._value, CellRichText):
                logger.debug(f"Found rich text in _value")
                rich_text = cell._value
                return self._parse_rich_text_object(rich_text, cell.coordinate, merged_info)
            
            # Method 2: Check value attribute
            if isinstance(cell.value, CellRichText):
                logger.debug(f"Found rich text in value")
                rich_text = cell.value
                return self._parse_rich_text_object(rich_text, cell.coordinate, merged_info)
            
            # Method 3: For merged cells, check the range's first cell
            if merged_info:
                try:
                    worksheet = cell.parent
                    top_left_cell = worksheet[merged_info['top_left']]
                    
                    # Check merged cell's main cell for rich text
                    if hasattr(top_left_cell, '_value') and isinstance(top_left_cell._value, CellRichText):
                        logger.debug(f"Found rich text in merged cell main cell")
                        rich_text = top_left_cell._value
                        return self._parse_rich_text_object(rich_text, cell.coordinate, merged_info)
                    elif isinstance(top_left_cell.value, CellRichText):
                        logger.debug(f"Found rich text in merged cell main cell value")
                        rich_text = top_left_cell.value
                        return self._parse_rich_text_object(rich_text, cell.coordinate, merged_info)
                except Exception as merged_err:
                    logger.debug(f"Error checking merged cell main cell: {merged_err}")
            
            # Method 4: Check for rich text attributes
            if hasattr(cell, 'richText') and cell.richText:
                logger.debug(f"Found traditional richText format")
                # Handle traditional richText format if needed
                return None
            
            # Method 5: Check raw data structure
            if hasattr(cell, '_value') and hasattr(cell._value, '__dict__'):
                logger.debug(f"_value attributes: {cell._value.__dict__}")
            
            # Method 6: Check all attributes for rich text
            rich_attrs = [attr for attr in dir(cell) if 'rich' in attr.lower()]
            if rich_attrs:
                logger.debug(f"Found rich text related attributes: {rich_attrs}")
                for attr in rich_attrs:
                    try:
                        value = getattr(cell, attr)
                        if value:
                            logger.debug(f"{attr}: {type(value)} = {value}")
                    except Exception:
                        pass
            
            logger.debug(f"No rich text format detected")
            return None
            
        except Exception as e:
            logger.error(f"Error extracting rich text format: {e}")
            return None

    def _parse_rich_text_object(
        self, rich_text: CellRichText, coordinate: str, merged_info: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Parse rich text object and extract formatting information.

        Args:
            rich_text: CellRichText object
            coordinate: Cell coordinate
            merged_info: Merged cell information

        Returns:
            Rich text information dictionary
        """
        rich_info = {
            "has_rich_text": True, 
            "segments": [],
            "merged_info": merged_info
        }

        logger.debug(f"Found rich text format in {coordinate}")
        if merged_info:
            logger.debug(f"Merged cell range: {merged_info['range']}")

        try:
            for i, item in enumerate(rich_text):
                if isinstance(item, TextBlock):
                    segment_info = {"text": item.text, "font": None, "segment_index": i}

                    # Extract font information
                    if item.font:
                        font_info = {
                            "name": getattr(item.font, "rFont", None),
                            "size": getattr(item.font, "sz", None),
                            "bold": getattr(item.font, "b", None),
                            "italic": getattr(item.font, "i", None),
                            "underline": getattr(item.font, "u", None),
                            "color": self._safe_copy_color(getattr(item.font, "color", None)) if getattr(item.font, "color", None) else None
                        }
                        
                        # Enhanced color information extraction
                        if getattr(item.font, "color", None):
                            font_color = getattr(item.font, "color", None)
                            font_info['color_raw'] = font_color
                            if hasattr(font_color, 'rgb') and font_color.rgb:
                                font_info['color_rgb'] = font_color.rgb
                            if hasattr(font_color, 'indexed') and font_color.indexed is not None:
                                font_info['color_indexed'] = font_color.indexed
                            if hasattr(font_color, 'theme') and font_color.theme is not None:
                                font_info['color_theme'] = font_color.theme
                                if hasattr(font_color, 'tint') and font_color.tint is not None:
                                    font_info['color_tint'] = font_color.tint
                        
                        segment_info["font"] = font_info
                        
                        # Debug color information
                        color_str = ""
                        if getattr(item.font, "color", None):
                            font_color = getattr(item.font, "color", None)
                            if hasattr(font_color, 'rgb') and font_color.rgb:
                                color_str = f" Color:#{font_color.rgb}"
                            elif hasattr(font_color, 'indexed') and font_color.indexed is not None:
                                color_str = f" Color:Index({font_color.indexed})"
                            elif hasattr(font_color, 'theme') and font_color.theme is not None:
                                color_str = f" Color:Theme({font_color.theme})"
                                if hasattr(font_color, 'tint') and font_color.tint is not None:
                                    color_str += f" Tint({font_color.tint})"
                            else:
                                color_str = " Color:present"
                        
                        logger.debug(f"Text segment {i}: '{item.text[:20]}...' {color_str}")
                    else:
                        logger.debug(f"Text segment {i}: '{item.text[:20]}...' no font")
                    
                    rich_info["segments"].append(segment_info)

                elif isinstance(item, str):
                    # Plain text segment
                    rich_info["segments"].append(
                        {"text": item, "font": None, "segment_index": i}
                    )
                    logger.debug(f"Plain text segment {i}: '{item[:20]}...'")

        except Exception as e:
            logger.error(f"Error parsing rich text object: {e}")
            import traceback
            traceback.print_exc()

        return rich_info

    def _apply_cell_format(
        self, cell, format_info: Dict[str, Any], target_language: str = "en"
    ) -> None:
        """
        Apply formatting to a cell.

        Args:
            cell: openpyxl cell object
            format_info: Format information dictionary
            target_language: Target language code
        """
        try:
            if not format_info:
                return

            # Apply font formatting
            font_kwargs = {}

            # Font name (with language-specific adjustments)
            if target_language == "th" and format_info.get("font_name"):
                # Use Thai-compatible font
                thai_fonts = ["TH SarabunPSK", "Tahoma", "Arial Unicode MS"]
                original_font = format_info["font_name"]
                if not any(
                    thai_font.lower() in original_font.lower()
                    for thai_font in thai_fonts
                ):
                    font_kwargs["name"] = "TH SarabunPSK"
                else:
                    font_kwargs["name"] = original_font
            elif format_info.get("font_name"):
                font_kwargs["name"] = format_info["font_name"]

            # Font size (with adjustment)
            if format_info.get("font_size"):
                original_size = format_info["font_size"]
                adjusted_size = max(6, int(original_size * self.font_size_adjustment))
                font_kwargs["size"] = adjusted_size

            # Other font properties
            for prop in ["font_bold", "font_italic", "font_underline", "font_strike"]:
                if format_info.get(prop) is not None:
                    font_kwargs[prop.replace("font_", "")] = format_info[prop]

            # Font color - create a new Color object to avoid StyleProxy issues
            if format_info.get("font_color"):
                try:
                    color_obj = format_info["font_color"]
                    if hasattr(color_obj, "rgb") and color_obj.rgb:
                        font_kwargs["color"] = Color(rgb=color_obj.rgb)
                    elif (
                        hasattr(color_obj, "indexed") and color_obj.indexed is not None
                    ):
                        font_kwargs["color"] = Color(indexed=color_obj.indexed)
                    elif hasattr(color_obj, "theme") and color_obj.theme is not None:
                        tint = getattr(color_obj, "tint", None)
                        font_kwargs["color"] = Color(theme=color_obj.theme, tint=tint)
                except Exception as e:
                    logger.debug(f"Could not apply font color: {e}")

            if font_kwargs:
                cell.font = Font(**font_kwargs)

            # Apply fill formatting - create a new PatternFill to avoid StyleProxy issues
            if format_info.get("fill_object"):
                try:
                    fill_obj = format_info["fill_object"]
                    if hasattr(fill_obj, "fill_type") and hasattr(
                        fill_obj, "start_color"
                    ):
                        start_color = fill_obj.start_color
                        if hasattr(start_color, "rgb") and start_color.rgb:
                            cell.fill = PatternFill(
                                fill_type=fill_obj.fill_type,
                                start_color=Color(rgb=start_color.rgb),
                            )
                        elif (
                            hasattr(start_color, "indexed")
                            and start_color.indexed is not None
                        ):
                            cell.fill = PatternFill(
                                fill_type=fill_obj.fill_type,
                                start_color=Color(indexed=start_color.indexed),
                            )
                        elif (
                            hasattr(start_color, "theme")
                            and start_color.theme is not None
                        ):
                            tint = getattr(start_color, "tint", None)
                            cell.fill = PatternFill(
                                fill_type=fill_obj.fill_type,
                                start_color=Color(theme=start_color.theme, tint=tint),
                            )
                except Exception as e:
                    logger.debug(f"Could not apply fill formatting: {e}")

            # Apply alignment
            alignment_kwargs = {}
            for prop in ["horizontal", "vertical", "wrap_text", "shrink_to_fit"]:
                if format_info.get(prop) is not None:
                    alignment_kwargs[prop] = format_info[prop]

            if alignment_kwargs:
                cell.alignment = Alignment(**alignment_kwargs)

            # Apply border - create new Border object to avoid StyleProxy issues
            if format_info.get("border"):
                try:
                    # For now, skip border application to avoid StyleProxy issues
                    # A more complete implementation would recreate the border object
                    pass
                except Exception as e:
                    logger.debug(f"Could not apply border: {e}")

            # Apply number format
            if format_info.get("number_format"):
                cell.number_format = format_info["number_format"]

        except Exception as e:
            logger.error(f"Error applying cell format: {e}")

    def _apply_rich_text_format(
        self,
        cell,
        original_text: str,
        translated_text: str,
        rich_text_info: Dict[str, Any],
        target_language: str = "en",
    ) -> None:
        """
        Apply rich text formatting to translated text with enhanced multi-segment support.

        Args:
            cell: openpyxl cell object
            original_text: Original text
            translated_text: Translated text
            rich_text_info: Rich text formatting information
            target_language: Target language code
        """
        if not rich_text_info or not rich_text_info.get("has_rich_text"):
            return

        try:
            logger.debug(f"Applying rich text format to {cell.coordinate}")
            
            segments = rich_text_info.get("segments", [])
            merged_info = rich_text_info.get("merged_info")
            
            if not segments:
                return

            # Handle merged cells specially
            target_cells = [cell]  # Default to just current cell
            
            if merged_info:
                logger.debug(f"Processing merged cell: {merged_info['range']}")
                # For merged cells, need to sync to all cells
                target_cells = merged_info.get('all_cells', [cell])
                logger.debug(f"Target cells count: {len(target_cells)}")

            # Create new rich text parts
            rich_text_parts = []

            # If only one segment, apply to entire translated text
            if len(segments) == 1:
                segment = segments[0]
                if segment.get("font"):
                    # Create inline font with language support
                    font_info = segment['font'].copy()
                    if target_language == 'th':
                        font_info['target_language'] = 'th'
                    inline_font = self._create_inline_font(font_info, target_language)
                    rich_text_parts.append(TextBlock(inline_font, translated_text))
                    logger.debug(f"Single segment applied: {segment.get('font', {}).get('color_rgb', 'default')}")
                else:
                    rich_text_parts.append(translated_text)
            else:
                # Multiple segments: use enhanced distribution algorithm
                self._distribute_translated_text_for_merged_cells(segments, original_text, translated_text, rich_text_parts, merged_info, target_language)

            # Apply rich text to all target cells
            if rich_text_parts:
                successful_cells = []
                failed_cells = []
                
                for target_cell in target_cells:
                    try:
                        target_cell._value = CellRichText(rich_text_parts)
                        successful_cells.append(target_cell.coordinate)
                    except Exception as apply_err:
                        logger.warning(f"Apply to {target_cell.coordinate} failed: {apply_err}")
                        failed_cells.append(target_cell.coordinate)
                        # Fall back to plain text
                        try:
                            target_cell.value = translated_text
                        except Exception:
                            pass
                
                if successful_cells:
                    logger.debug(f"Rich text applied successfully to: {', '.join(successful_cells)}")
                if failed_cells:
                    logger.warning(f"Application failed for cells: {', '.join(failed_cells)}")

        except Exception as e:
            logger.error(f"Error applying rich text format: {e}")
            import traceback
            traceback.print_exc()
            # Fall back to plain text
            cell.value = translated_text
    
    def _distribute_translated_text_for_merged_cells(self, segments: List[Dict], original_text: str, 
                                                    translated_text: str, rich_text_parts: List, 
                                                    merged_info: Optional[Dict[str, Any]], target_language: str = 'en') -> None:
        """
        Enhanced text distribution algorithm optimized for merged cells.
        
        Args:
            segments: Original text segments list
            original_text: Original complete text
            translated_text: Translated complete text
            rich_text_parts: Rich text parts list (output)
            merged_info: Merged cell information
            target_language: Target language code
        """
        try:
            logger.debug(f"Enhanced text distribution for merged cells")
            if merged_info:
                logger.debug(f"Merged range: {merged_info.get('range', 'unknown')}")
            
            # For merged cells, use more intelligent distribution strategy
            if len(segments) <= 2:
                # If few segments, distribute by proportion
                self._distribute_translated_text(segments, original_text, translated_text, rich_text_parts, target_language)
                return
            
            # For multi-segment merged cells, prioritize main color segments
            # Find the longest segment as main segment
            main_segment = max(segments, key=lambda s: len(s.get('text', '')))
            main_segment_index = segments.index(main_segment)
            
            # Distribution strategy: main segment gets 70% of translated text, others get remaining
            main_portion = 0.7
            
            translated_len = len(translated_text)
            main_text_len = int(translated_len * main_portion)
            other_text_len = translated_len - main_text_len
            
            # Distribute text
            other_segments = [s for i, s in enumerate(segments) if i != main_segment_index]
            other_segment_len = other_text_len // len(other_segments) if other_segments else 0
            
            current_pos = 0
            for i, segment in enumerate(segments):
                if i == main_segment_index:
                    # Main segment
                    segment_text = translated_text[current_pos:current_pos + main_text_len]
                    current_pos += main_text_len
                else:
                    # Other segments
                    if i == len(segments) - 1:
                        # Last segment, use all remaining text
                        segment_text = translated_text[current_pos:]
                    else:
                        segment_text = translated_text[current_pos:current_pos + other_segment_len]
                        current_pos += other_segment_len
                
                # Create text block with language support
                if segment.get("font"):
                    font_info = segment['font'].copy()
                    if target_language == 'th':
                        font_info['target_language'] = 'th'
                    inline_font = self._create_inline_font(font_info, target_language)
                    rich_text_parts.append(TextBlock(inline_font, segment_text))
                    
                    # Display color info
                    color_info = ""
                    if segment.get('font', {}).get('color_rgb'):
                        color_info = f" Color:#{segment['font']['color_rgb']}"
                    elif segment.get('font', {}).get('color_indexed'):
                        color_info = f" Color:Indexed({segment['font']['color_indexed']})"
                    elif segment.get('font', {}).get('color_theme'):
                        color_info = f" Color:Theme({segment['font']['color_theme']})"
                    
                    logger.debug(f"Segment {i}: '{segment_text[:20]}...'{color_info}")
                else:
                    rich_text_parts.append(segment_text)
                    logger.debug(f"Segment {i}: '{segment_text[:20]}...' no format")
            
        except Exception as e:
            logger.warning(f"Enhanced text distribution failed: {e}")
            # Fall back to simple distribution
            self._distribute_translated_text(segments, original_text, translated_text, rich_text_parts, target_language)
    
    def _distribute_translated_text(self, segments: List[Dict], original_text: str, 
                                   translated_text: str, rich_text_parts: List, target_language: str = 'en') -> None:
        """
        Distribute translated text proportionally among segments.
        
        Args:
            segments: Original text segments list
            original_text: Original complete text
            translated_text: Translated complete text
            rich_text_parts: Rich text parts list (output)
            target_language: Target language code
        """
        try:
            # Calculate proportions for each segment
            total_length = len(original_text)
            if total_length == 0:
                return
            
            # Simplification: if too many segments, use first segment's format for entire text
            if len(segments) > 5:
                first_segment = segments[0]
                if first_segment.get("font"):
                    font_info = first_segment['font'].copy()
                    if target_language == 'th':
                        font_info['target_language'] = 'th'
                    inline_font = self._create_inline_font(font_info, target_language)
                    rich_text_parts.append(TextBlock(inline_font, translated_text))
                else:
                    rich_text_parts.append(translated_text)
                return
            
            # Proportional distribution
            translated_pos = 0
            for i, segment in enumerate(segments):
                segment_text = segment.get('text', '')
                segment_length = len(segment_text)
                
                if segment_length == 0:
                    continue
                
                # Calculate proportion for this segment
                if i == len(segments) - 1:
                    # Last segment, use all remaining text
                    segment_translated = translated_text[translated_pos:]
                else:
                    # Proportional calculation
                    proportion = segment_length / total_length
                    segment_translated_length = int(len(translated_text) * proportion)
                    segment_translated = translated_text[translated_pos:translated_pos + segment_translated_length]
                    translated_pos += segment_translated_length
                
                # Create text block
                if segment.get("font"):
                    font_info = segment['font'].copy()
                    if target_language == 'th':
                        font_info['target_language'] = 'th'
                    inline_font = self._create_inline_font(font_info, target_language)
                    rich_text_parts.append(TextBlock(inline_font, segment_translated))
                else:
                    rich_text_parts.append(segment_translated)
            
        except Exception as e:
            logger.warning(f"Text distribution failed: {e}")
            # Fall back: use first segment's format
            if segments:
                first_segment = segments[0]
                if first_segment.get("font"):
                    font_info = first_segment['font'].copy()
                    if target_language == 'th':
                        font_info['target_language'] = 'th'
                    inline_font = self._create_inline_font(font_info, target_language)
                    rich_text_parts.append(TextBlock(inline_font, translated_text))
                else:
                    rich_text_parts.append(translated_text)

    def _create_inline_font(
        self, font_info: Dict[str, Any], target_language: str = "en"
    ) -> InlineFont:
        """
        Create an InlineFont object from font information with enhanced color support.

        Args:
            font_info: Font information dictionary
            target_language: Target language code

        Returns:
            InlineFont object
        """
        font_kwargs = {}

        if font_info.get("name"):
            font_kwargs["rFont"] = font_info["name"]
        # For Thai rich text, set appropriate font
        elif 'target_language' in font_info and font_info['target_language'] == 'th':
            font_kwargs["rFont"] = "TH SarabunPSK"
        elif target_language == "th":
            font_kwargs["rFont"] = "TH SarabunPSK"

        if font_info.get("size"):
            font_kwargs["sz"] = font_info["size"]
        if font_info.get("bold"):
            font_kwargs["b"] = font_info["bold"]
        if font_info.get("italic"):
            font_kwargs["i"] = font_info["italic"]
        if font_info.get("underline"):
            # Fix underline value validation issue
            underline_value = font_info["underline"]
            if underline_value is True:
                font_kwargs["u"] = "single"
            elif underline_value in ['single', 'singleAccounting', 'double', 'doubleAccounting']:
                font_kwargs["u"] = underline_value
            # Other cases don't set underline

        # Enhanced color handling
        if font_info.get("color"):
            try:
                font_kwargs["color"] = font_info["color"]
                logger.debug(f"Using original color object")
            except Exception as color_err:
                logger.debug(f"Using original color object failed: {color_err}")
                
                # Try using backup color information
                if font_info.get('color_rgb'):
                    try:
                        font_kwargs["color"] = Color(rgb=font_info['color_rgb'])
                        logger.debug(f"Using RGB color: #{font_info['color_rgb']}")
                    except Exception as rgb_err:
                        logger.debug(f"Using RGB color failed: {rgb_err}")
                        
                elif font_info.get('color_indexed') is not None:
                    try:
                        font_kwargs["color"] = Color(indexed=font_info['color_indexed'])
                        logger.debug(f"Using indexed color: {font_info['color_indexed']}")
                    except Exception as idx_err:
                        logger.debug(f"Using indexed color failed: {idx_err}")
                        
                elif font_info.get('color_theme') is not None:
                    try:
                        if font_info.get('color_tint') is not None:
                            font_kwargs["color"] = Color(theme=font_info['color_theme'], tint=font_info['color_tint'])
                            logger.debug(f"Using theme color: {font_info['color_theme']} tint: {font_info['color_tint']}")
                        else:
                            font_kwargs["color"] = Color(theme=font_info['color_theme'])
                            logger.debug(f"Using theme color: {font_info['color_theme']}")
                    except Exception as theme_err:
                        logger.debug(f"Using theme color failed: {theme_err}")

        return InlineFont(**font_kwargs)
    
    def _safe_copy_color(self, color_obj) -> Optional[Color]:
        """
        Safely copy color object to avoid StyleProxy issues.
        
        Args:
            color_obj: Original color object
            
        Returns:
            New color object or None
        """
        if not color_obj:
            return None
        
        try:
            # Method 1: Priority use RGB values
            if hasattr(color_obj, 'rgb') and color_obj.rgb:
                new_color = Color(rgb=color_obj.rgb)
                logger.debug(f"Copied RGB color: #{color_obj.rgb}")
                return new_color
            
            # Method 2: Use indexed color
            elif hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
                new_color = Color(indexed=color_obj.indexed)
                logger.debug(f"Copied indexed color: {color_obj.indexed}")
                return new_color
            
            # Method 3: Use theme color
            elif hasattr(color_obj, 'theme') and color_obj.theme is not None:
                if hasattr(color_obj, 'tint') and color_obj.tint is not None:
                    new_color = Color(theme=color_obj.theme, tint=color_obj.tint)
                    logger.debug(f"Copied theme color: {color_obj.theme} tint: {color_obj.tint}")
                else:
                    new_color = Color(theme=color_obj.theme)
                    logger.debug(f"Copied theme color: {color_obj.theme}")
                return new_color
            
            # Method 4: Use auto color
            elif hasattr(color_obj, 'auto') and color_obj.auto is not None:
                new_color = Color(auto=color_obj.auto)
                logger.debug(f"Copied auto color: {color_obj.auto}")
                return new_color
            
            # Method 5: Try to return original object
            else:
                logger.debug(f"Using original color object")
                return color_obj
                
        except Exception as e:
            logger.warning(f"Failed to copy color object: {e}")
            
            # Final fallback: try to extract all possible color information
            try:
                # Check object attributes
                if hasattr(color_obj, '__dict__'):
                    attrs = color_obj.__dict__
                    logger.debug(f"Color object attributes: {attrs}")
                    
                    # Try to construct new color object
                    color_kwargs = {}
                    for attr in ['rgb', 'indexed', 'theme', 'tint', 'auto']:
                        if hasattr(color_obj, attr) and getattr(color_obj, attr) is not None:
                            color_kwargs[attr] = getattr(color_obj, attr)
                    
                    if color_kwargs:
                        new_color = Color(**color_kwargs)
                        logger.debug(f"Constructed color from attributes: {color_kwargs}")
                        return new_color
                
                # If all methods fail, return original object
                return color_obj
                
            except Exception as backup_err:
                logger.warning(f"Backup color copy method also failed: {backup_err}")
                return color_obj  # Return original object as last resort
    
    def _check_merged_cell(self, cell) -> Optional[Dict[str, Any]]:
        """
        Check if cell is part of a merged cell and return related information.
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            Merged cell information dictionary or None
        """
        try:
            worksheet = cell.parent
            if not worksheet or not hasattr(worksheet, 'merged_cells'):
                return None
            
            cell_coord = cell.coordinate
            for merged_range in worksheet.merged_cells.ranges:
                if cell_coord in merged_range:
                    # Get all cells in the merged range
                    all_cells = []
                    for row in worksheet[merged_range.coord]:
                        if isinstance(row, (list, tuple)):
                            all_cells.extend(row)
                        else:
                            all_cells.append(row)
                    
                    return {
                        'is_merged': True,
                        'range': str(merged_range),
                        'top_left': merged_range.coord.split(':')[0],
                        'bottom_right': merged_range.coord.split(':')[1] if ':' in merged_range.coord else merged_range.coord.split(':')[0],
                        'all_cells': all_cells,
                        'merged_range_obj': merged_range
                    }
            
            return None
            
        except Exception as e:
            logger.error(f"Error checking merged cell: {e}")
            return None
    
    def _synchronize_merged_cell_formats(self, cell, original_text: str, translated_text: str, 
                                        format_info: Dict[str, Any], rich_text_info: Optional[Dict[str, Any]], 
                                        merged_cell_info: Dict[str, Any]) -> None:
        """
        Synchronize merged cell formats to all related cells.
        
        Args:
            cell: Current cell
            original_text: Original text
            translated_text: Translated text
            format_info: Format information
            rich_text_info: Rich text information
            merged_cell_info: Merged cell information
        """
        try:
            logger.debug(f"Synchronizing merged cell format: {merged_cell_info['range']}")
            
            # Get all merged cells
            all_cells = merged_cell_info.get('all_cells', [])
            if not all_cells:
                logger.debug(f"No merged cell list found, using backup method")
                # Backup method: manually get from worksheet
                worksheet = cell.parent
                merged_range = merged_cell_info['merged_range_obj']
                for row_cells in worksheet[merged_range.coord]:
                    if isinstance(row_cells, (list, tuple)):
                        all_cells.extend(row_cells)
                    else:
                        all_cells.append(row_cells)
            
            # Synchronize to all cells
            successful_syncs = []
            failed_syncs = []
            
            for target_cell in all_cells:
                try:
                    # Skip current cell (already processed)
                    if target_cell.coordinate == cell.coordinate:
                        continue
                        
                    # Set text value first
                    target_cell.value = translated_text
                    
                    # Apply basic format
                    if format_info:
                        self._apply_cell_format(target_cell, format_info)
                    
                    # Apply rich text format if available
                    if rich_text_info and rich_text_info.get("has_rich_text"):
                        self._apply_rich_text_format(
                            target_cell, original_text, translated_text, rich_text_info
                        )
                    
                    successful_syncs.append(target_cell.coordinate)
                    
                except Exception as sync_err:
                    logger.warning(f"Synchronization to {target_cell.coordinate} failed: {sync_err}")
                    failed_syncs.append(target_cell.coordinate)
                    
                    # Try to at least sync text content
                    try:
                        target_cell.value = translated_text
                    except Exception:
                        pass
            
            # Report synchronization results
            if successful_syncs:
                logger.debug(f"Successfully synchronized to: {', '.join(successful_syncs)}")
            if failed_syncs:
                logger.warning(f"Synchronization failed: {', '.join(failed_syncs)}")
            
            # Special handling: if rich text exists and there are failures, try simpler sync
            if rich_text_info and failed_syncs:
                logger.debug(f"Trying simplified synchronization method...")
                for coord in failed_syncs:
                    try:
                        target_cell = cell.parent[coord]
                        # Use first segment's format for entire text
                        segments = rich_text_info.get('segments', [])
                        if segments and segments[0].get('font'):
                            inline_font = self._create_inline_font(segments[0]['font'])
                            target_cell._value = CellRichText([TextBlock(inline_font, translated_text)])
                            logger.debug(f"Simplified sync successful: {coord}")
                    except Exception as simple_err:
                        logger.warning(f"Simplified sync also failed: {coord} - {simple_err}")
            
        except Exception as e:
            logger.error(f"Error synchronizing merged cell formats: {e}")

    def _smart_adjust_column_width(self, workbook) -> None:
        """
        Intelligently adjust column widths to fit content.

        Args:
            workbook: openpyxl workbook object
        """
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Get image-occupied columns if image protection is enabled
                occupied_columns = set()
                if self.image_protection and sheet_name in self.image_data:
                    for img_info in self.image_data[sheet_name]:
                        anchor_info = img_info.get("anchor_info", {})
                        if anchor_info.get("type") == "two_cell":
                            from_col = anchor_info.get("from_col", 0)
                            to_col = anchor_info.get("to_col", 0)
                            for col in range(from_col, to_col + 1):
                                occupied_columns.add(col)

                # Adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    column_index = column[0].column

                    # Check if column has images
                    is_occupied = column_index in occupied_columns

                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except Exception:
                            pass

                    # Set column width (conservative for image-occupied columns)
                    if is_occupied:
                        adjusted_width = min(max_length + 1, 30)
                    else:
                        adjusted_width = min(max_length + 2, 50)

                    sheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logger.error(f"Error adjusting column widths: {e}")


# Backward compatibility alias
ExcelTranslator = ExcelProcessor
