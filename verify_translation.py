#!/usr/bin/env python3
"""
Verify the translation results
"""

from openpyxl import load_workbook

def verify_translation_results():
    """Verify the translation results"""
    print("=== Verifying Translation Results ===\n")
    
    # Check Chinese translation
    print("1. Checking Chinese Translation:")
    print("-" * 40)
    
    try:
        wb_zh = load_workbook("/root/projects/github/Offitrans/test_output.xlsx", data_only=False)
        sheet_zh = wb_zh["Translation Test"]
        
        original_expected = [
            ("A1", "Translation Service Test Excel"),
            ("A2", "This Excel file contains English content for translation testing."),
            ("A4", "Name"),
            ("B4", "Country"), 
            ("C4", "Language"),
            ("A5", "Alice"),
            ("B5", "USA"),
            ("C5", "English"),
            ("A6", "Bob"),
            ("B6", "France"),
            ("C6", "French")
        ]
        
        for cell_addr, original_text in original_expected:
            cell = sheet_zh[cell_addr]
            translated = str(cell.value) if cell.value else ""
            
            # Check if text was translated (different from original)
            was_translated = translated != original_text and translated.strip() != ""
            status = "✅ Translated" if was_translated else "❌ Not translated"
            
            print(f"  {cell_addr}: '{original_text}' → '{translated}' {status}")
        
        wb_zh.close()
        
    except Exception as e:
        print(f"Error checking Chinese translation: {e}")
    
    # Check Thai translation
    print(f"\n2. Checking Thai Translation:")
    print("-" * 40)
    
    try:
        wb_th = load_workbook("/root/projects/github/Offitrans/test_output_th.xlsx", data_only=False)
        sheet_th = wb_th["Translation Test"]
        
        for cell_addr, original_text in original_expected:
            cell = sheet_th[cell_addr]
            translated = str(cell.value) if cell.value else ""
            
            # Check if text was translated (different from original)
            was_translated = translated != original_text and translated.strip() != ""
            status = "✅ Translated" if was_translated else "❌ Not translated"
            
            print(f"  {cell_addr}: '{original_text}' → '{translated}' {status}")
        
        wb_th.close()
        
    except Exception as e:
        print(f"Error checking Thai translation: {e}")

if __name__ == "__main__":
    verify_translation_results()