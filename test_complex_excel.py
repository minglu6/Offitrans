#!/usr/bin/env python3
"""
Create a complex Excel file to test comprehensive text extraction
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import logging

logging.basicConfig(level=logging.INFO)

def create_complex_excel():
    """Create a complex Excel file with various formatting scenarios"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Complex Test"
    
    # 1. Regular text
    ws['A1'] = "Regular Text"
    
    # 2. Text with special formatting
    ws['A2'] = "Bold Italic Text"
    ws['A2'].font = Font(bold=True, italic=True, color=Color(rgb="FF0000"))
    
    # 3. Rich text with multiple colors (this should test our rich text extraction)
    rich_text = CellRichText([
        TextBlock(InlineFont(rFont="Arial", sz=12, color=Color(rgb="FF0000")), "红色文本 "),
        TextBlock(InlineFont(rFont="Arial", sz=12, color=Color(rgb="0000FF")), "蓝色文本 "),
        TextBlock(InlineFont(rFont="Arial", sz=12, color=Color(rgb="00FF00")), "绿色文本")
    ])
    ws['A3'] = rich_text
    
    # 4. Merged cells
    ws.merge_cells('A4:C4')
    ws['A4'] = "合并单元格中的文本"
    ws['A4'].alignment = Alignment(horizontal='center')
    
    # 5. Formula (should be skipped)
    ws['A5'] = "=SUM(1+1)"
    
    # 6. Number that should be skipped
    ws['A6'] = 123.45
    
    # 7. Text that looks like number but is string
    ws['A7'] = "123.45 元"
    
    # 8. Multi-line text
    ws['A8'] = "第一行\n第二行\n第三行"
    
    # 9. Empty cell (should be skipped)
    # A9 is empty
    
    # 10. Text with special characters
    ws['A10'] = "特殊字符：!@#$%^&*()_+-={}[]|\\:;\"'<>?,./"
    
    # 11. Another rich text with merged cells
    ws.merge_cells('B2:D2')
    rich_text2 = CellRichText([
        TextBlock(InlineFont(rFont="Arial", sz=14, b=True, color=Color(rgb="800080")), "紫色粗体 "),
        TextBlock(InlineFont(rFont="Arial", sz=10, i=True, color=Color(rgb="FFA500")), "橙色斜体")
    ])
    ws['B2'] = rich_text2
    
    # 12. Text in row 78 column M-Q (specifically mentioned in reference code)
    if ws.max_row < 78:
        # Add some placeholder rows
        for i in range(ws.max_row + 1, 78):
            ws[f'A{i}'] = f"占位文本第{i}行"
    
    # Add text to row 78, columns M-Q
    ws['M78'] = "第78行M列文本"
    ws['N78'] = "第78行N列文本"
    ws['O78'] = "第78行O列文本"
    ws['P78'] = "第78行P列文本"
    ws['Q78'] = "第78行Q列文本"
    
    # Make some of them rich text
    rich_text_78 = CellRichText([
        TextBlock(InlineFont(rFont="Arial", sz=12, color=Color(rgb="FF1493")), "粉红色 "),
        TextBlock(InlineFont(rFont="Arial", sz=12, color=Color(rgb="4169E1")), "蓝色")
    ])
    ws['N78'] = rich_text_78
    
    # Save the file
    filename = "/root/projects/github/Offitrans/complex_test.xlsx"
    wb.save(filename)
    print(f"Created complex Excel file: {filename}")
    
    return filename

if __name__ == "__main__":
    complex_file = create_complex_excel()
    
    # Now test it with our analyzer
    print("\n" + "="*60)
    print("Testing complex Excel file...")
    print("="*60)
    
    # Import our debug function
    import sys
    sys.path.append('/root/projects/github/Offitrans')
    from debug_excel import analyze_excel_file
    
    analyze_excel_file(complex_file)