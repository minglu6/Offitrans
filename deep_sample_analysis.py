#!/usr/bin/env python3
"""
Deep analysis of sample.xlsx to ensure we're not missing anything
"""

from openpyxl import load_workbook
import logging

def deep_analyze_sample():
    """Deep analysis of the sample file"""
    file_path = "/root/projects/github/Offitrans/examples/sample_files/sample.xlsx"
    print(f"=== Deep Analysis of: {file_path} ===\n")
    
    # Load with different parameters
    print("1. Loading with data_only=False (formulas preserved):")
    wb1 = load_workbook(file_path, data_only=False)
    analyze_workbook(wb1, "data_only=False")
    wb1.close()
    
    print("\n2. Loading with data_only=True (formulas evaluated):")
    wb2 = load_workbook(file_path, data_only=True)
    analyze_workbook(wb2, "data_only=True")
    wb2.close()
    
    print("\n3. Loading with keep_vba=True:")
    try:
        wb3 = load_workbook(file_path, keep_vba=True)
        analyze_workbook(wb3, "keep_vba=True")
        wb3.close()
    except Exception as e:
        print(f"keep_vba failed: {e}")

def analyze_workbook(workbook, mode):
    """Analyze a workbook"""
    print(f"\n--- Analysis Mode: {mode} ---")
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"  Max row: {sheet.max_row}, Max column: {sheet.max_column}")
        print(f"  Used range: A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}")
        
        # Check every single cell in the used range
        text_cells = []
        formula_cells = []
        numeric_cells = []
        empty_cells = []
        other_cells = []
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                
                if cell.value is None:
                    empty_cells.append(cell.coordinate)
                elif isinstance(cell.value, str):
                    if cell.value.startswith('='):
                        formula_cells.append((cell.coordinate, cell.value))
                    else:
                        text_cells.append((cell.coordinate, cell.value))
                elif isinstance(cell.value, (int, float)):
                    numeric_cells.append((cell.coordinate, cell.value))
                else:
                    other_cells.append((cell.coordinate, type(cell.value), cell.value))
        
        print(f"  Text cells: {len(text_cells)}")
        for coord, value in text_cells:
            print(f"    {coord}: '{value}'")
        
        print(f"  Formula cells: {len(formula_cells)}")
        for coord, value in formula_cells:
            print(f"    {coord}: '{value}'")
            
        print(f"  Numeric cells: {len(numeric_cells)}")
        for coord, value in numeric_cells:
            print(f"    {coord}: {value}")
            
        print(f"  Empty cells: {len(empty_cells)}")
        
        if other_cells:
            print(f"  Other cell types: {len(other_cells)}")
            for coord, cell_type, value in other_cells:
                print(f"    {coord}: {cell_type} = {value}")
        
        # Check for merged cells
        if hasattr(sheet, 'merged_cells') and sheet.merged_cells:
            print(f"  Merged cells: {len(sheet.merged_cells.ranges)}")
            for merged_range in sheet.merged_cells.ranges:
                print(f"    {merged_range}")
        
        # Check for images
        if hasattr(sheet, '_images') and sheet._images:
            print(f"  Images: {len(sheet._images)}")

if __name__ == "__main__":
    deep_analyze_sample()