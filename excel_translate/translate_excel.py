#!/usr/bin/env python3
"""
Excel翻译工具 V2 - 解决图片变形问题
支持将Excel文件中的文字翻译成指定语言，同时保持原有格式和图片不变形

主要改进：
1. 添加图片信息提取和保存功能
2. 在翻译过程中保持图片原始位置和尺寸
3. 优化单元格尺寸调整逻辑
4. 增强错误处理机制

依赖库：
pip install openpyxl requests pillow
"""

import os
import time
from typing import List, Dict, Optional, Any
import re
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
import io
try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    print("⚠️ PIL(Pillow)库未安装，图片验证功能将被禁用")
    print("ℹ️ 可以使用 'pip install pillow' 安装")


class ExcelTranslatorV2:
    def __init__(self, translate_api_key: Optional[str] = None, font_size_adjustment: float = 0.8):
        """
        初始化Excel翻译器 V2
        
        Args:
            translate_api_key: Google翻译API密钥
            font_size_adjustment: 字体大小调整比例（默认0.8，即缩小到80%）
        """
        self.translate_api_key = translate_api_key
        self.font_size_adjustment = font_size_adjustment
        self.image_data: Dict[str, List[Dict[str, Any]]] = {}  # 存储图片信息
        self.image_fallback_enabled = True  # 启用图片备选处理
        self.translation_stats = {}  # 翻译统计信息
        
    def _safe_copy_color(self, color_obj) -> Optional[Color]:
        """
        安全复制颜色对象
        
        Args:
            color_obj: 原始颜色对象
            
        Returns:
            新的颜色对象或None
        """
        if not color_obj:
            return None
        
        try:
            # 方法1：优先使用RGB值
            if hasattr(color_obj, 'rgb') and color_obj.rgb:
                new_color = Color(rgb=color_obj.rgb)
                print(f"    🎨 复制RGB颜色: #{color_obj.rgb}")
                return new_color
            
            # 方法2：使用索引颜色
            elif hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
                new_color = Color(indexed=color_obj.indexed)
                print(f"    🎨 复制索引颜色: {color_obj.indexed}")
                return new_color
            
            # 方法3：使用主题颜色
            elif hasattr(color_obj, 'theme') and color_obj.theme is not None:
                if hasattr(color_obj, 'tint') and color_obj.tint is not None:
                    new_color = Color(theme=color_obj.theme, tint=color_obj.tint)
                    print(f"    🎨 复制主题颜色: {color_obj.theme} tint: {color_obj.tint}")
                else:
                    new_color = Color(theme=color_obj.theme)
                    print(f"    🎨 复制主题颜色: {color_obj.theme}")
                return new_color
            
            # 方法4：使用自动颜色
            elif hasattr(color_obj, 'auto') and color_obj.auto is not None:
                new_color = Color(auto=color_obj.auto)
                print(f"    🎨 复制自动颜色: {color_obj.auto}")
                return new_color
            
            # 方法5：尝试直接返回原始对象
            else:
                print(f"    🎨 使用原始颜色对象")
                return color_obj
                
        except Exception as e:
            print(f"⚠️ 复制颜色对象失败: {e}")
            
            # 最后的备选方案：尝试从原始对象中提取所有可能的颜色信息
            try:
                # 检查对象的所有属性
                if hasattr(color_obj, '__dict__'):
                    attrs = color_obj.__dict__
                    print(f"    🔍 颜色对象属性: {attrs}")
                    
                    # 尝试构造新的颜色对象
                    color_kwargs = {}
                    for attr in ['rgb', 'indexed', 'theme', 'tint', 'auto']:
                        if hasattr(color_obj, attr) and getattr(color_obj, attr) is not None:
                            color_kwargs[attr] = getattr(color_obj, attr)
                    
                    if color_kwargs:
                        new_color = Color(**color_kwargs)
                        print(f"    🎨 通过属性构造颜色: {color_kwargs}")
                        return new_color
                
                # 如果所有方法都失败，返回原始对象
                return color_obj
                
            except Exception as backup_err:
                print(f"⚠️ 备选颜色复制方案也失败: {backup_err}")
                return color_obj  # 返回原始对象作为最后的备选
    
    def _check_merged_cell(self, cell) -> Optional[Dict[str, Any]]:
        """
        检查单元格是否为合并单元格并返回相关信息
        
        Args:
            cell: openpyxl单元格对象
            
        Returns:
            合并单元格信息字典或None
        """
        try:
            worksheet = cell.parent
            if not worksheet or not hasattr(worksheet, 'merged_cells'):
                return None
            
            cell_coord = cell.coordinate
            for merged_range in worksheet.merged_cells.ranges:
                if cell_coord in merged_range:
                    # 获取合并单元格范围的所有单元格
                    all_cells = []
                    for row in worksheet[merged_range.coord]:
                        if isinstance(row, (list, tuple)):
                            all_cells.extend(row)
                        else:
                            all_cells.append(row)
                    
                    return {
                        'is_merged': True,
                        'range': str(merged_range),
                        'top_left': merged_range.coord.split(':')[0],
                        'bottom_right': merged_range.coord.split(':')[1] if ':' in merged_range.coord else merged_range.coord.split(':')[0],
                        'all_cells': all_cells,
                        'merged_range_obj': merged_range
                    }
            
            return None
            
        except Exception as e:
            print(f"⚠️ 检查合并单元格时出错: {e}")
            return None
    
    def _extract_rich_text_format(self, cell) -> Optional[Dict[str, Any]]:
        """
        提取单元格内的富文本格式信息
        
        Args:
            cell: openpyxl单元格对象
            
        Returns:
            富文本格式信息字典或None
        """
        try:
            # 增强调试信息
            cell_text = str(cell.value) if cell.value else ""
            print(f"  🔍 检查单元格 {cell.coordinate}: '{cell_text[:30]}...'")
            print(f"    - 单元格类型: {type(cell.value)}")
            print(f"    - _value类型: {type(cell._value) if hasattr(cell, '_value') else 'None'}")
            
            # 检查合并单元格状态
            merged_info = None
            if hasattr(cell, 'coordinate'):
                worksheet = cell.parent
                if worksheet and hasattr(worksheet, 'merged_cells'):
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            print(f"    - ⚙️ 检测到合并单元格: {merged_range}")
                            merged_info = {
                                'range': str(merged_range),
                                'top_left': merged_range.coord.split(':')[0]
                            }
                            break
            
            # 方法1: 检查_value属性
            if hasattr(cell, '_value') and isinstance(cell._value, CellRichText):
                print(f"    - ✅ 发现_value中的富文本")
                rich_text = cell._value
                return self._parse_rich_text_object(rich_text, cell.coordinate, merged_info)
            
            # 方法2: 检查value属性
            if isinstance(cell.value, CellRichText):
                print(f"    - ✅ 发现value中的富文本")
                rich_text = cell.value
                return self._parse_rich_text_object(rich_text, cell.coordinate, merged_info)
            
            # 方法3: 对于合并单元格，检查范围内的第一个单元格
            if merged_info:
                try:
                    worksheet = cell.parent
                    top_left_cell = worksheet[merged_info['top_left']]
                    
                    # 检查合并单元格的主单元格是否有富文本
                    if hasattr(top_left_cell, '_value') and isinstance(top_left_cell._value, CellRichText):
                        print(f"    - ✅ 在合并单元格主单元格中发现富文本")
                        rich_text = top_left_cell._value
                        return self._parse_rich_text_object(rich_text, cell.coordinate, merged_info)
                    elif isinstance(top_left_cell.value, CellRichText):
                        print(f"    - ✅ 在合并单元格主单元格value中发现富文本")
                        rich_text = top_left_cell.value
                        return self._parse_rich_text_object(rich_text, cell.coordinate, merged_info)
                except Exception as merged_err:
                    print(f"    - ⚠️ 检查合并单元格主单元格时出错: {merged_err}")
            
            # 方法4: 检查是否有丰富文本属性
            if hasattr(cell, 'richText') and cell.richText:
                print(f"    - ✅ 发现传统richText格式")
                # 这里可以处理传统的richText格式
                return None
            
            # 方法5: 检查原始数据结构
            if hasattr(cell, '_value') and hasattr(cell._value, '__dict__'):
                print(f"    - 🔍 _value属性: {cell._value.__dict__}")
            
            # 方法6: 检查所有属性
            rich_attrs = [attr for attr in dir(cell) if 'rich' in attr.lower()]
            if rich_attrs:
                print(f"    - 🔍 发现富文本相关属性: {rich_attrs}")
                for attr in rich_attrs:
                    try:
                        value = getattr(cell, attr)
                        if value:
                            print(f"      - {attr}: {type(value)} = {value}")
                    except Exception:
                        pass
            
            print(f"    - ❌ 未检测到富文本格式")
            return None
            
        except Exception as e:
            print(f"⚠️ 提取富文本格式时出错: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _parse_rich_text_object(self, rich_text: CellRichText, coordinate: str, merged_info: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        解析富文本对象
        
        Args:
            rich_text: CellRichText对象
            coordinate: 单元格坐标
            merged_info: 合并单元格信息
            
        Returns:
            富文本信息字典
        """
        rich_info = {
            'has_rich_text': True,
            'segments': [],
            'merged_info': merged_info
        }
        
        print(f"  🎨 发现富文本格式: {coordinate}")
        if merged_info:
            print(f"    📎 合并单元格范围: {merged_info['range']}")
        
        try:
            # 遍历富文本段落
            for i, item in enumerate(rich_text):
                if isinstance(item, TextBlock):
                    segment_info = {
                        'text': item.text,
                        'font': None,
                        'segment_index': i
                    }
                    
                    # 提取字体信息
                    if item.font:
                        font_info = {
                            'name': getattr(item.font, 'rFont', None),
                            'size': getattr(item.font, 'sz', None),
                            'bold': getattr(item.font, 'b', None),
                            'italic': getattr(item.font, 'i', None),
                            'underline': getattr(item.font, 'u', None),
                            'color': self._safe_copy_color(getattr(item.font, 'color', None)) if getattr(item.font, 'color', None) else None
                        }
                        
                        # 增强颜色信息提取
                        if getattr(item.font, 'color', None):
                            font_color = getattr(item.font, 'color', None)
                            font_info['color_raw'] = font_color
                            if hasattr(font_color, 'rgb') and font_color.rgb:
                                font_info['color_rgb'] = font_color.rgb
                            if hasattr(font_color, 'indexed') and font_color.indexed is not None:
                                font_info['color_indexed'] = font_color.indexed
                            if hasattr(font_color, 'theme') and font_color.theme is not None:
                                font_info['color_theme'] = font_color.theme
                                if hasattr(font_color, 'tint') and font_color.tint is not None:
                                    font_info['color_tint'] = font_color.tint
                        
                        segment_info['font'] = font_info
                        
                        # 调试信息
                        color_str = ""
                        if getattr(item.font, 'color', None):
                            font_color = getattr(item.font, 'color', None)
                            if hasattr(font_color, 'rgb') and font_color.rgb:
                                color_str = f" 颜色:#{font_color.rgb}"
                            elif hasattr(font_color, 'indexed') and font_color.indexed is not None:
                                color_str = f" 颜色:Index({font_color.indexed})"
                            elif hasattr(font_color, 'theme') and font_color.theme is not None:
                                color_str = f" 颜色:Theme({font_color.theme})"
                                if hasattr(font_color, 'tint') and font_color.tint is not None:
                                    color_str += f" Tint({font_color.tint})"
                            else:
                                color_str = " 颜色:有"
                        
                        print(f"    📝 文本段{i}: '{item.text[:20]}...' {color_str}")
                    else:
                        print(f"    📝 文本段{i}: '{item.text[:20]}...' 无字体")
                    
                    rich_info['segments'].append(segment_info)
                elif isinstance(item, str):
                    # 纯文本段落
                    rich_info['segments'].append({
                        'text': item,
                        'font': None,
                        'segment_index': i
                    })
                    print(f"    📝 纯文本段{i}: '{item[:20]}...'")
        
        except Exception as e:
            print(f"⚠️ 解析富文本时出错: {e}")
            import traceback
            traceback.print_exc()
        
        return rich_info
    
    def _apply_rich_text_format(self, cell, original_text: str, translated_text: str, 
                               rich_text_info: Optional[Dict[str, Any]], target_language: str = 'th') -> None:
        """
        应用富文本格式到翻译后的文本（优化合并单元格支持）
        
        Args:
            cell: openpyxl单元格对象
            original_text: 原始文本
            translated_text: 翻译后的文本
            rich_text_info: 富文本格式信息
            target_language: 目标语言代码
        """
        if not rich_text_info or not rich_text_info.get('has_rich_text'):
            return
        
        try:
            print(f"  🎨 应用富文本格式到 {cell.coordinate}")
            
            segments = rich_text_info.get('segments', [])
            merged_info = rich_text_info.get('merged_info')
            
            if not segments:
                return
            
            # 如果是合并单元格，特殊处理
            target_cells = [cell]  # 默认只处理当前单元格
            
            if merged_info:
                print(f"    📎 处理合并单元格: {merged_info['range']}")
                # 获取合并单元格范围的所有单元格
                worksheet = cell.parent
                top_left_cell = worksheet[merged_info['top_left']]
                
                # 对于合并单元格，需要同步更新所有单元格
                target_cells = merged_info.get('all_cells', [cell])
                print(f"    🎯 目标单元格数量: {len(target_cells)}")
            
            # 创建新的富文本对象
            rich_text_parts = []
            
            # 如果只有一个段落，直接应用到整个翻译文本
            if len(segments) == 1:
                segment = segments[0]
                if segment.get('font'):
                    # 创建内联字体（支持泰文字体）
                    font_info = segment['font'].copy()
                    if target_language == 'th':
                        font_info['target_language'] = 'th'
                    inline_font = self._create_inline_font(font_info)
                    rich_text_parts.append(TextBlock(inline_font, translated_text))
                    print(f"    ✅ 单段落应用: {segment.get('font', {}).get('color_rgb', 'default')}")
                else:
                    rich_text_parts.append(translated_text)
            else:
                # 多个段落：优化分配算法
                self._distribute_translated_text_for_merged_cells(segments, original_text, translated_text, rich_text_parts, merged_info, target_language)
            
            # 应用富文本到所有目标单元格
            if rich_text_parts:
                successful_cells = []
                failed_cells = []
                
                for target_cell in target_cells:
                    try:
                        target_cell._value = CellRichText(rich_text_parts)
                        successful_cells.append(target_cell.coordinate)
                    except Exception as apply_err:
                        print(f"    ⚠️ 应用到 {target_cell.coordinate} 失败: {apply_err}")
                        failed_cells.append(target_cell.coordinate)
                        # 回退到普通文本
                        try:
                            target_cell.value = translated_text
                        except Exception:
                            pass
                
                if successful_cells:
                    print(f"    ✅ 富文本格式应用成功到: {', '.join(successful_cells)}")
                if failed_cells:
                    print(f"    ⚠️ 应用失败的单元格: {', '.join(failed_cells)}")
            
        except Exception as e:
            print(f"⚠️ 应用富文本格式时出错: {e}")
            import traceback
            traceback.print_exc()
            # 如果富文本应用失败，回退到普通文本
            cell.value = translated_text
    
    def _create_inline_font(self, font_info: Dict[str, Any]) -> InlineFont:
        """
        创建内联字体对象
        
        Args:
            font_info: 字体信息字典
            
        Returns:
            InlineFont对象
        """
        font_kwargs = {}
        
        if font_info.get('name'):
            font_kwargs['rFont'] = font_info['name']
        # 为泰文富文本设置合适的字体
        elif 'target_language' in font_info and font_info['target_language'] == 'th':
            font_kwargs['rFont'] = 'TH SarabunPSK'
        if font_info.get('size'):
            font_kwargs['sz'] = font_info['size']
        if font_info.get('bold'):
            font_kwargs['b'] = font_info['bold']
        if font_info.get('italic'):
            font_kwargs['i'] = font_info['italic']
        if font_info.get('underline'):
            # 修复下划线的值验证问题
            underline_value = font_info['underline']
            if underline_value is True:
                font_kwargs['u'] = 'single'
            elif underline_value in ['single', 'singleAccounting', 'double', 'doubleAccounting']:
                font_kwargs['u'] = underline_value
            # 其他情况不设置下划线
        
        # 增强颜色处理
        if font_info.get('color'):
            try:
                font_kwargs['color'] = font_info['color']
                print(f"      🎨 使用原始颜色对象")
            except Exception as color_err:
                print(f"      ⚠️ 使用原始颜色对象失败: {color_err}")
                
                # 尝试使用备用颜色信息
                if font_info.get('color_rgb'):
                    try:
                        font_kwargs['color'] = Color(rgb=font_info['color_rgb'])
                        print(f"      🎨 使用RGB颜色: #{font_info['color_rgb']}")
                    except Exception as rgb_err:
                        print(f"      ⚠️ 使用RGB颜色失败: {rgb_err}")
                        
                elif font_info.get('color_indexed') is not None:
                    try:
                        font_kwargs['color'] = Color(indexed=font_info['color_indexed'])
                        print(f"      🎨 使用索引颜色: {font_info['color_indexed']}")
                    except Exception as idx_err:
                        print(f"      ⚠️ 使用索引颜色失败: {idx_err}")
                        
                elif font_info.get('color_theme') is not None:
                    try:
                        if font_info.get('color_tint') is not None:
                            font_kwargs['color'] = Color(theme=font_info['color_theme'], tint=font_info['color_tint'])
                            print(f"      🎨 使用主题颜色: {font_info['color_theme']} tint: {font_info['color_tint']}")
                        else:
                            font_kwargs['color'] = Color(theme=font_info['color_theme'])
                            print(f"      🎨 使用主题颜色: {font_info['color_theme']}")
                    except Exception as theme_err:
                        print(f"      ⚠️ 使用主题颜色失败: {theme_err}")
        
        return InlineFont(**font_kwargs)
    
    def _distribute_translated_text(self, segments: List[Dict], original_text: str, 
                                   translated_text: str, rich_text_parts: List) -> None:
        """
        将翻译后的文本按比例分配给不同的格式段落
        
        Args:
            segments: 原始文本段落列表
            original_text: 原始完整文本
            translated_text: 翻译后的完整文本
            rich_text_parts: 富文本部分列表（输出）
        """
        try:
            # 计算每个段落的长度比例
            total_length = len(original_text)
            if total_length == 0:
                return
            
            # 简化处理：如果段落太多，使用第一个段落的格式应用到整个翻译文本
            if len(segments) > 5:
                first_segment = segments[0]
                if first_segment.get('font'):
                    inline_font = self._create_inline_font(first_segment['font'])
                    rich_text_parts.append(TextBlock(inline_font, translated_text))
                else:
                    rich_text_parts.append(translated_text)
                return
            
            # 按比例分配翻译文本
            translated_pos = 0
            for i, segment in enumerate(segments):
                segment_text = segment.get('text', '')
                segment_length = len(segment_text)
                
                if segment_length == 0:
                    continue
                
                # 计算这个段落应该占翻译文本的比例
                if i == len(segments) - 1:
                    # 最后一个段落，使用剩余的所有文本
                    segment_translated = translated_text[translated_pos:]
                else:
                    # 按比例计算
                    proportion = segment_length / total_length
                    segment_translated_length = int(len(translated_text) * proportion)
                    segment_translated = translated_text[translated_pos:translated_pos + segment_translated_length]
                    translated_pos += segment_translated_length
                
                # 创建文本块
                if segment.get('font'):
                    inline_font = self._create_inline_font(segment['font'])
                    rich_text_parts.append(TextBlock(inline_font, segment_translated))
                else:
                    rich_text_parts.append(segment_translated)
            
        except Exception as e:
            print(f"⚠️ 分配翻译文本时出错: {e}")
            # 回退：使用第一个段落的格式
            if segments:
                first_segment = segments[0]
                if first_segment.get('font'):
                    inline_font = self._create_inline_font(first_segment['font'])
                    rich_text_parts.append(TextBlock(inline_font, translated_text))
                else:
                    rich_text_parts.append(translated_text)
    
    def _distribute_translated_text_for_merged_cells(self, segments: List[Dict], original_text: str, 
                                                    translated_text: str, rich_text_parts: List, 
                                                    merged_info: Optional[Dict[str, Any]], target_language: str = 'th') -> None:
        """
        为合并单元格优化的文本分配算法
        
        Args:
            segments: 原始文本段落列表
            original_text: 原始完整文本
            translated_text: 翻译后的完整文本
            rich_text_parts: 富文本部分列表（输出）
            merged_info: 合并单元格信息
            target_language: 目标语言代码
        """
        try:
            print(f"    🔄 为合并单元格优化文本分配")
            if merged_info:
                print(f"    📎 合并范围: {merged_info.get('range', 'unknown')}")
            
            # 对于合并单元格，使用更智能的分配策略
            if len(segments) <= 2:
                # 如果段落少，直接按比例分配
                self._distribute_translated_text(segments, original_text, translated_text, rich_text_parts)
                return
            
            # 对于多段落的合并单元格，优先保持主要颜色段落
            # 找到最长的段落作为主要段落
            main_segment = max(segments, key=lambda s: len(s.get('text', '')))
            main_segment_index = segments.index(main_segment)
            
            # 分配策略：主要段落占大部分翻译文本，其他段落占小部分
            main_portion = 0.7  # 主要段落占总长度的70%
            
            translated_len = len(translated_text)
            main_text_len = int(translated_len * main_portion)
            other_text_len = translated_len - main_text_len
            
            # 分配文本
            other_segments = [s for i, s in enumerate(segments) if i != main_segment_index]
            other_segment_len = other_text_len // len(other_segments) if other_segments else 0
            
            current_pos = 0
            for i, segment in enumerate(segments):
                if i == main_segment_index:
                    # 主要段落
                    segment_text = translated_text[current_pos:current_pos + main_text_len]
                    current_pos += main_text_len
                else:
                    # 其他段落
                    if i == len(segments) - 1:
                        # 最后一个段落，使用剩余所有文本
                        segment_text = translated_text[current_pos:]
                    else:
                        segment_text = translated_text[current_pos:current_pos + other_segment_len]
                        current_pos += other_segment_len
                
                # 创建文本块（支持泰文字体）
                if segment.get('font'):
                    font_info = segment['font'].copy()
                    if target_language == 'th':
                        font_info['target_language'] = 'th'
                    inline_font = self._create_inline_font(font_info)
                    rich_text_parts.append(TextBlock(inline_font, segment_text))
                    
                    # 显示颜色信息
                    color_info = ""
                    if segment.get('font', {}).get('color_rgb'):
                        color_info = f" 颜色:#{segment['font']['color_rgb']}"
                    elif segment.get('font', {}).get('color_indexed'):
                        color_info = f" 颜色:Indexed({segment['font']['color_indexed']})"
                    elif segment.get('font', {}).get('color_theme'):
                        color_info = f" 颜色:Theme({segment['font']['color_theme']})"
                    
                    print(f"      ✅ 段落{i}: '{segment_text[:20]}...'{color_info}")
                else:
                    rich_text_parts.append(segment_text)
                    print(f"      ✅ 段落{i}: '{segment_text[:20]}...' 无格式")
            
        except Exception as e:
            print(f"⚠️ 合并单元格文本分配时出错: {e}")
            # 回退到普通分配策略
            self._distribute_translated_text(segments, original_text, translated_text, rich_text_parts)
    
    def _synchronize_merged_cell_formats(self, cell, original_text: str, translated_text: str, 
                                        format_info: Dict[str, Any], rich_text_info: Optional[Dict[str, Any]], 
                                        merged_cell_info: Dict[str, Any]) -> None:
        """
        同步合并单元格中的格式到所有相关单元格
        
        Args:
            cell: 当前单元格
            original_text: 原始文本
            translated_text: 翻译后的文本
            format_info: 格式信息
            rich_text_info: 富文本信息
            merged_cell_info: 合并单元格信息
        """
        try:
            print(f"    🔄 同步合并单元格格式: {merged_cell_info['range']}")
            
            # 获取所有合并单元格
            all_cells = merged_cell_info.get('all_cells', [])
            if not all_cells:
                print(f"    ⚠️ 未找到合并单元格列表，使用备用方法")
                # 备用方法：从工作表中手动获取
                worksheet = cell.parent
                merged_range = merged_cell_info['merged_range_obj']
                for row_cells in worksheet[merged_range.coord]:
                    if isinstance(row_cells, (list, tuple)):
                        all_cells.extend(row_cells)
                    else:
                        all_cells.append(row_cells)
            
            # 同步到所有单元格
            successful_syncs = []
            failed_syncs = []
            
            for target_cell in all_cells:
                try:
                    # 跳过当前单元格（已经处理过）
                    if target_cell.coordinate == cell.coordinate:
                        continue
                        
                    # 先设置文本值
                    target_cell.value = translated_text
                    
                    # 应用基本格式
                    if format_info:
                        self._apply_cell_format(target_cell, format_info)
                    
                    # 应用富文本格式（如果有）
                    if rich_text_info and rich_text_info.get('has_rich_text'):
                        self._apply_rich_text_format(target_cell, original_text, translated_text, rich_text_info)
                    
                    successful_syncs.append(target_cell.coordinate)
                    
                except Exception as sync_err:
                    print(f"    ⚠️ 同步到 {target_cell.coordinate} 失败: {sync_err}")
                    failed_syncs.append(target_cell.coordinate)
                    
                    # 尝试至少同步文本内容
                    try:
                        target_cell.value = translated_text
                    except Exception:
                        pass
            
            # 报告同步结果
            if successful_syncs:
                print(f"    ✅ 成功同步到: {', '.join(successful_syncs)}")
            if failed_syncs:
                print(f"    ⚠️ 同步失败: {', '.join(failed_syncs)}")
            
            # 特别处理：如果有富文本且存在失败，尝试更简单的同步方式
            if rich_text_info and failed_syncs:
                print(f"    🔄 尝试简化同步方式...")
                for coord in failed_syncs:
                    try:
                        target_cell = cell.parent[coord]
                        # 使用第一个段落的格式应用到整个文本
                        segments = rich_text_info.get('segments', [])
                        if segments and segments[0].get('font'):
                            inline_font = self._create_inline_font(segments[0]['font'])
                            target_cell._value = CellRichText([TextBlock(inline_font, translated_text)])
                            print(f"    ✅ 简化同步成功: {coord}")
                    except Exception as simple_err:
                        print(f"    ⚠️ 简化同步也失败: {coord} - {simple_err}")
            
        except Exception as e:
            print(f"    ⚠️ 同步合并单元格格式时出错: {e}")
            import traceback
            traceback.print_exc()
    
    def _safe_create_image(self, img_obj) -> Optional[Image]:
        """
        安全创建图片对象，处理各种可能的错误
        
        Args:
            img_obj: 原始图片对象
            
        Returns:
            新的图片对象或None
        """
        try:
            # 方法1: 直接使用原始对象（最安全）
            if hasattr(img_obj, 'anchor'):
                print("✅ 使用原始图片对象（推荐方式）")
                return img_obj
            
            # 方法2: 尝试使用_data()方法
            if hasattr(img_obj, '_data'):
                try:
                    img_data = img_obj._data()
                    if img_data:
                        # 检查和清理数据
                        if isinstance(img_data, bytes):
                            # 移除null字节
                            if b'\x00' in img_data:
                                print("⚠️ 检测到null字节，正在清理...")
                                img_data = img_data.replace(b'\x00', b'')
                            
                            # 验证图片数据（如果PIL可用）
                            if PIL_AVAILABLE:
                                try:
                                    # 使用PIL验证图片数据
                                    test_img = PILImage.open(io.BytesIO(img_data))
                                    test_img.verify()
                                    print("✅ 图片数据验证成功")
                                except Exception as pil_err:
                                    print(f"⚠️ PIL验证失败: {pil_err}")
                                    # 继续尝试使用数据
                            else:
                                print("ℹ️ 跳过PIL验证（未安装）")
                            
                            # 创建新的openpyxl图片对象
                            try:
                                new_img = Image(img_data)
                                print("✅ 使用清理后的数据创建图片成功")
                                return new_img
                            except Exception as create_err:
                                print(f"⚠️ 使用清理后数据创建图片失败: {create_err}")
                                pass
                        
                except Exception as data_err:
                    print(f"⚠️ 获取图片数据失败: {data_err}")
            
            # 方法3: 尝试使用其他属性
            if hasattr(img_obj, 'ref'):
                try:
                    print("ℹ️ 尝试使用图片引用")
                    # 这里可能需要从工作簿中重新加载图片
                    return img_obj
                except Exception:
                    pass
            
            # 如果所有方法都失败，返回原始对象
            print("⚠️ 所有方法都失败，返回原始对象")
            return img_obj
            
        except Exception as e:
            print(f"❌ 创建图片对象完全失败: {e}")
            return None
        
    def extract_images_info(self, workbook) -> Dict[str, List[Dict[str, Any]]]:
        """
        提取Excel中的图片信息
        
        Args:
            workbook: openpyxl工作簿对象
            
        Returns:
            图片信息字典
        """
        images_info = {}
        
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_images = []
                
                # 检查是否有图片
                if hasattr(sheet, '_images') and sheet._images:
                    print(f"📷 在工作表 '{sheet_name}' 中找到 {len(sheet._images)} 个图片")
                    
                    for img in sheet._images:
                        img_info = {
                            'image_object': img,
                            'anchor_type': type(img.anchor).__name__,
                        }
                        
                        # 提取锚点信息
                        if isinstance(img.anchor, TwoCellAnchor):
                            img_info['anchor_info'] = {
                                'type': 'two_cell',
                                'from_col': img.anchor._from.col,
                                'from_col_off': img.anchor._from.colOff,
                                'from_row': img.anchor._from.row,
                                'from_row_off': img.anchor._from.rowOff,
                                'to_col': img.anchor.to.col,
                                'to_col_off': img.anchor.to.colOff,
                                'to_row': img.anchor.to.row,
                                'to_row_off': img.anchor.to.rowOff,
                            }
                        elif isinstance(img.anchor, OneCellAnchor):
                            img_info['anchor_info'] = {
                                'type': 'one_cell',
                                'from_col': img.anchor._from.col,
                                'from_col_off': img.anchor._from.colOff,
                                'from_row': img.anchor._from.row,
                                'from_row_off': img.anchor._from.rowOff,
                                'width': img.anchor.ext.cx,
                                'height': img.anchor.ext.cy,
                            }
                        
                        sheet_images.append(img_info)
                
                images_info[sheet_name] = sheet_images
                
        except Exception as e:
            print(f"⚠️ 提取图片信息时出错: {e}")
            
        return images_info
    
    def restore_images_info(self, workbook, images_info: Dict[str, List[Dict[str, Any]]]) -> None:
        """
        恢复Excel中的图片信息
        
        Args:
            workbook: openpyxl工作簿对象
            images_info: 图片信息字典
        """
        try:
            for sheet_name, sheet_images in images_info.items():
                if not sheet_images:
                    continue
                    
                sheet = workbook[sheet_name]
                
                # 清除现有图片（如果有的话）
                if hasattr(sheet, '_images'):
                    sheet._images.clear()
                else:
                    sheet._images = []
                
                # 恢复图片
                for img_info in sheet_images:
                    try:
                        img_obj = img_info['image_object']
                        
                        # 使用安全的图片创建方法
                        new_img = self._safe_create_image(img_obj)
                        if new_img is None:
                            print("❌ 无法创建图片对象，跳过此图片")
                            continue
                        
                        # 恢复锚点信息
                        anchor_info = img_info.get('anchor_info', {})
                        if anchor_info.get('type') == 'two_cell':
                            # 创建TwoCellAnchor
                            from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
                            
                            anchor = TwoCellAnchor()
                            anchor._from.col = anchor_info['from_col']
                            anchor._from.colOff = anchor_info['from_col_off']
                            anchor._from.row = anchor_info['from_row'] 
                            anchor._from.rowOff = anchor_info['from_row_off']
                            anchor.to.col = anchor_info['to_col']
                            anchor.to.colOff = anchor_info['to_col_off']
                            anchor.to.row = anchor_info['to_row']
                            anchor.to.rowOff = anchor_info['to_row_off']
                            
                        elif anchor_info.get('type') == 'one_cell':
                            # 创建OneCellAnchor
                            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
                            
                            anchor = OneCellAnchor()
                            anchor._from.col = anchor_info['from_col']
                            anchor._from.colOff = anchor_info['from_col_off']
                            anchor._from.row = anchor_info['from_row']
                            anchor._from.rowOff = anchor_info['from_row_off']
                            anchor.ext.cx = anchor_info['width']
                            anchor.ext.cy = anchor_info['height']
                        else:
                            # 使用原始锚点
                            anchor = img_obj.anchor
                        
                        new_img.anchor = anchor
                        try:
                            sheet.add_image(new_img)
                            print(f"✅ 成功添加图片到工作表 {sheet_name}")
                        except Exception as add_err:
                            print(f"⚠️ 添加图片到工作表失败: {add_err}")
                            # 尝试使用默认锚点重新添加
                            try:
                                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
                                default_anchor = OneCellAnchor()
                                new_img.anchor = default_anchor
                                sheet.add_image(new_img)
                                print("✅ 使用默认锚点成功添加图片")
                            except Exception as default_err:
                                print(f"❌ 使用默认锚点也失败: {default_err}")
                                continue
                        
                    except Exception as e:
                        print(f"⚠️ 恢复图片时出错: {e}")
                        # 如果无法恢复锚点，尝试替代方案
                        try:
                            print("🔄 尝试使用原始图片对象...")
                            # 检查原始图片对象的状态
                            if hasattr(img_obj, 'anchor') and img_obj.anchor:
                                sheet.add_image(img_obj)
                                print("✅ 使用原始图片对象成功")
                            else:
                                # 创建一个简单的默认锚点
                                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
                                default_anchor = OneCellAnchor()
                                default_anchor._from.col = 0
                                default_anchor._from.row = 0
                                default_anchor._from.colOff = 0
                                default_anchor._from.rowOff = 0
                                
                                # 设置默认大小
                                default_anchor.ext.cx = 2000000  # 默认宽度
                                default_anchor.ext.cy = 2000000  # 默认高度
                                
                                img_obj.anchor = default_anchor
                                sheet.add_image(img_obj)
                                print("✅ 使用默认锚点成功")
                        except Exception as fallback_err:
                            print(f"❌ 所有图片恢复方案都失败: {fallback_err}")
                            print("ℹ️ 跳过此图片，继续处理其他图片")
                            continue
                        
        except Exception as e:
            print(f"⚠️ 恢复图片信息时出错: {e}")

    def extract_text_from_excel(self, excel_path: str) -> List[Dict[str, Any]]:
        """
        从Excel文件中提取所有文本内容
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            包含文本信息的列表
        """
        text_data = []
        
        try:
            workbook = load_workbook(excel_path, data_only=False)
            print(f"✅ 成功打开Excel文件: {excel_path}")
            
            # 提取图片信息
            print("📷 提取图片信息...")
            self.image_data = self.extract_images_info(workbook)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f"处理工作表: {sheet_name}")
                
                # 遍历所有单元格
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.strip():
                            # 跳过公式单元格（以=开头）
                            if not cell.value.startswith('='):
                                # 保存原始格式信息
                                format_info = self._extract_cell_format(cell)
                                
                                # 检查是否有富文本格式
                                rich_text_info = self._extract_rich_text_format(cell)
                                
                                text_data.append({
                                    'text': cell.value,
                                    'sheet_name': sheet_name,
                                    'row': cell.row,
                                    'column': cell.column,
                                    'cell_coordinate': cell.coordinate,
                                    'format_info': format_info,
                                    'rich_text_info': rich_text_info
                                })
                                
                                # 构建颜色信息显示
                                color_info = ""
                                if format_info.get('font_color'):
                                    if format_info.get('font_color_rgb'):
                                        color_info = f" 颜色:#{format_info['font_color_rgb']}"
                                    elif format_info.get('font_color_indexed'):
                                        color_info = f" 颜色:Indexed({format_info['font_color_indexed']})"
                                    elif format_info.get('font_color_theme'):
                                        color_info = f" 颜色:Theme({format_info['font_color_theme']})"
                                    else:
                                        color_info = " 颜色:有"
                                
                                print(f"  提取文本: {sheet_name}!{cell.coordinate} = '{cell.value[:50]}...' " +
                                      f"(字体: {format_info.get('font_name', 'Unknown')} " +
                                      f"{format_info.get('font_size', 'Unknown')}pt{color_info})")
                                
                                # 特别检查第78行的M-Q列
                                if cell.row == 78 and cell.column >= 13 and cell.column <= 17:  # M=13, Q=17
                                    print(f"    🔴 特别关注: 第78行 M-Q列 {cell.coordinate}")
                                    print(f"      - 文本内容: '{cell.value}'")
                                    print(f"      - 富文本信息: {rich_text_info}")
                                    
                                    # 详细检查这个单元格
                                    print(f"      - 原始内容检查:")
                                    print(f"        * cell.value: {type(cell.value)} = {cell.value}")
                                    print(f"        * cell._value: {type(cell._value) if hasattr(cell, '_value') else 'None'}")
                                    
                                    # 检查合并单元格
                                    merged_info = self._check_merged_cell(cell)
                                    if merged_info:
                                        print(f"      - 合并单元格信息: {merged_info}")
            
            workbook.close()
            print(f"✅ 总共提取了 {len(text_data)} 个文本单元格")
            return text_data
            
        except Exception as e:
            print(f"❌ 提取文本时发生错误: {e}")
            return []

    def _extract_cell_format(self, cell) -> Dict[str, Any]:
        """
        提取单元格格式信息
        
        Args:
            cell: openpyxl单元格对象
            
        Returns:
            格式信息字典
        """
        format_info = {}
        
        try:
            # 字体信息
            if cell.font:
                format_info['font_name'] = cell.font.name
                format_info['font_size'] = cell.font.size
                format_info['font_bold'] = cell.font.bold
                format_info['font_italic'] = cell.font.italic
                format_info['font_underline'] = cell.font.underline
                format_info['font_strike'] = cell.font.strike
                
                # 改进颜色处理
                if cell.font.color:
                    try:
                        # 保存完整的颜色对象
                        format_info['font_color'] = cell.font.color
                        # 同时保存颜色值用于调试
                        if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                            format_info['font_color_rgb'] = cell.font.color.rgb
                        elif hasattr(cell.font.color, 'indexed') and cell.font.color.indexed is not None:
                            format_info['font_color_indexed'] = cell.font.color.indexed
                        elif hasattr(cell.font.color, 'theme') and cell.font.color.theme is not None:
                            format_info['font_color_theme'] = cell.font.color.theme
                            if hasattr(cell.font.color, 'tint') and cell.font.color.tint is not None:
                                format_info['font_color_tint'] = cell.font.color.tint
                    except Exception as color_err:
                        print(f"⚠️ 提取字体颜色时出错: {color_err}")
                        format_info['font_color'] = None
            
            # 填充信息
            if cell.fill and hasattr(cell.fill, 'start_color'):
                try:
                    format_info['fill_color'] = cell.fill.start_color
                    format_info['fill_type'] = cell.fill.fill_type
                    # 保存完整的填充对象
                    format_info['fill_object'] = cell.fill
                except Exception as fill_err:
                    print(f"⚠️ 提取填充信息时出错: {fill_err}")
            
            # 对齐信息
            if cell.alignment:
                format_info['horizontal'] = cell.alignment.horizontal
                format_info['vertical'] = cell.alignment.vertical
                format_info['wrap_text'] = cell.alignment.wrap_text
                format_info['shrink_to_fit'] = cell.alignment.shrink_to_fit
            
            # 边框信息
            if cell.border:
                format_info['has_border'] = True
                format_info['border'] = cell.border
            
            # 数字格式
            if cell.number_format:
                format_info['number_format'] = cell.number_format
                
        except Exception as e:
            print(f"⚠️ 提取格式信息时出错: {e}")
        
        return format_info

    def translate_text_google(self, text: str, target_language: str = 'th') -> str:
        """
        使用Google翻译API翻译文本（支持中文到泰文）
        
        Args:
            text: 要翻译的文本
            target_language: 目标语言代码 ('th'=泰文, 'en'=英文)
            
        Returns:
            翻译后的文本
        """
        try:
            # 使用Google翻译的免费API
            url = "https://translate.googleapis.com/translate_a/single"
            params = {
                'client': 'gtx',
                'sl': 'auto',  # 自动检测源语言
                'tl': target_language,
                'dt': 't',
                'q': text
            }
            
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            
            result = response.json()
            if result and len(result) > 0 and len(result[0]) > 0:
                translated_text = result[0][0][0]
                return translated_text
            else:
                print(f"⚠️ 翻译API返回空结果: {text}")
                return text
                
        except Exception as e:
            print(f"❌ 翻译失败: {e}")
            return text

    def _should_translate_text(self, text: str) -> bool:
        """
        判断文本是否需要翻译
        
        Args:
            text: 待检查的文本
            
        Returns:
            是否需要翻译
        """
        if not text or not text.strip():
            return False
        
        text = text.strip()
        
        # 过滤纯数字
        if text.replace('.', '').replace(',', '').replace('-', '').replace('+', '').replace('%', '').replace('$', '').replace('€', '').replace('￥', '').replace('(', '').replace(')', '').replace(' ', '').isdigit():
            return False
        
        # 过滤纯符号
        symbol_chars = set('!@#$%^&*()_+-=[]{}|;:,.<>?/~`')
        if all(c in symbol_chars or c.isspace() for c in text):
            return False
        
        # 过滤纯英文字母和数字的组合（可能是代码或ID）
        if text.replace(' ', '').replace('-', '').replace('_', '').isalnum() and not re.search(r'[\u4e00-\u9fff]', text):
            # 但保留可能是英文单词的情况
            words = re.findall(r'[a-zA-Z]+', text)
            if not words or all(len(word) <= 2 for word in words):
                return False
        
        # 过滤日期格式
        date_patterns = [
            r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$',  # 2023-12-31
            r'^\d{1,2}[-/]\d{1,2}[-/]\d{4}$',  # 31-12-2023
            r'^\d{4}年\d{1,2}月\d{1,2}日$',     # 2023年12月31日
        ]
        for pattern in date_patterns:
            if re.match(pattern, text):
                return False
        
        # 过滤时间格式
        time_patterns = [
            r'^\d{1,2}:\d{2}(:\d{2})?$',      # 14:30:00
            r'^\d{1,2}:\d{2}\s*(AM|PM|am|pm)$', # 2:30 PM
        ]
        for pattern in time_patterns:
            if re.match(pattern, text):
                return False
        
        # 过滤邮箱
        if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', text):
            return False
        
        # 过滤URL
        if re.match(r'^https?://[\w\.-]+', text):
            return False
        
        # 检查是否包含中文
        if re.search(r'[\u4e00-\u9fff]', text):
            return True
        
        # 检查是否是有意义的英文文本（包含常见英文单词）
        words = re.findall(r'[a-zA-Z]+', text)
        if words and any(len(word) > 3 for word in words):
            return True
        
        return False
    
    def _analyze_and_deduplicate_texts(self, texts: List[str]) -> Dict[str, Any]:
        """
        分析和去重文本
        
        Args:
            texts: 原始文本列表
            
        Returns:
            包含去重信息的字典
        """
        print("📊 分析和去重文本...")
        
        # 统计信息
        total_texts = len(texts)
        unique_texts = set()
        texts_to_translate = set()
        filtered_out = []
        
        # 分析每个文本
        for text in texts:
            if text and text.strip():
                text = text.strip()
                unique_texts.add(text)
                
                if self._should_translate_text(text):
                    texts_to_translate.add(text)
                else:
                    filtered_out.append(text)
        
        # 创建映射
        unique_list = list(texts_to_translate)
        
        # 打印统计信息
        print(f"📈 文本统计分析:")
        print(f"  - 总文本数: {total_texts}")
        print(f"  - 唯一文本数: {len(unique_texts)}")
        print(f"  - 需要翻译的文本数: {len(texts_to_translate)}")
        print(f"  - 过滤掉的文本数: {len(filtered_out)}")
        print(f"  - 重复率: {((total_texts - len(unique_texts)) / total_texts * 100):.1f}%")
        print(f"  - 过滤率: {(len(filtered_out) / len(unique_texts) * 100):.1f}%")
        
        if filtered_out:
            print(f"  - 过滤示例: {filtered_out[:5]}")
        
        return {
            'unique_texts': unique_list,
            'total_count': total_texts,
            'unique_count': len(unique_texts),
            'translate_count': len(texts_to_translate),
            'filtered_count': len(filtered_out),
            'filtered_examples': filtered_out[:10]
        }
    
    def translate_text_batch(self, texts: List[str], target_language: str = 'th') -> List[str]:
        """
        批量翻译文本（优化版：去重、过滤、统计）
        
        Args:
            texts: 要翻译的文本列表
            target_language: 目标语言代码 ('th'=泰文, 'en'=英文)
            
        Returns:
            翻译后的文本列表
        """
        if not texts:
            return []
        
        # 第1步：分析和去重
        analysis = self._analyze_and_deduplicate_texts(texts)
        unique_texts = analysis['unique_texts']
        
        if not unique_texts:
            print("⚠️ 没有需要翻译的文本")
            return texts
        
        # 第2步：翻译去重后的文本
        print(f"🚀 开始翻译 {len(unique_texts)} 个唯一文本...")
        unique_translations = {}
        
        for i, text in enumerate(unique_texts):
            print(f"翻译进度: {i+1}/{len(unique_texts)} - {text[:30]}...")
            
            try:
                translated = self.translate_text_google(text, target_language)
                unique_translations[text] = translated
                print(f"  ✅ 翻译结果: {translated[:50]}{'...' if len(translated) > 50 else ''}")
            except Exception as e:
                print(f"  ❌ 翻译失败: {e}")
                unique_translations[text] = text  # 保持原文
            
            # 添加延迟避免API限制
            time.sleep(0.5)
        
        # 第3步：映射回原始文本列表
        print("📝 映射翻译结果...")
        translated_texts = []
        
        for original_text in texts:
            if not original_text or not original_text.strip():
                translated_texts.append(original_text)
                continue
            
            text = original_text.strip()
            
            # 检查是否需要翻译
            if self._should_translate_text(text):
                # 使用翻译结果
                translated = unique_translations.get(text, text)
                translated_texts.append(translated)
            else:
                # 保持原文
                translated_texts.append(original_text)
        
        # 打印最终统计
        print(f"📊 翻译完成统计:")
        print(f"  - 实际翻译API调用次数: {len(unique_texts)}")
        print(f"  - 节省API调用次数: {len(texts) - len(unique_texts)}")
        print(f"  - API调用优化率: {((len(texts) - len(unique_texts)) / len(texts) * 100):.1f}%")
        
        return translated_texts

    def replace_text_in_excel(self, excel_path: str, output_path: str, 
                             target_language: str = 'th') -> bool:
        """
        翻译Excel文件中的文本并保持格式和图片
        
        Args:
            excel_path: 输入Excel文件路径
            output_path: 输出Excel文件路径
            target_language: 目标语言代码 ('th'=泰文, 'en'=英文)
            
        Returns:
            是否成功
        """
        try:
            # 1. 提取所有文本
            print("📋 第1步: 提取Excel中的文本和图片...")
            text_data = self.extract_text_from_excel(excel_path)
            
            if not text_data:
                print("❌ 未找到可翻译的文本")
                return False
            
            # 2. 批量翻译
            print("🌍 第2步: 批量翻译文本...")
            original_texts = [item['text'] for item in text_data]
            translated_texts = self.translate_text_batch(original_texts, target_language)
            
            # 3. 替换文本并保持格式和图片
            print("📝 第3步: 替换文本并保持格式和图片...")
            success = self._replace_text_with_format_and_images(excel_path, output_path, 
                                                               text_data, translated_texts, target_language)
            
            if success:
                print(f"✅ 翻译完成！输出文件: {output_path}")
                return True
            else:
                print("❌ 替换文本时发生错误")
                return False
                
        except Exception as e:
            print(f"❌ 翻译Excel时发生错误: {e}")
            return False

    def _replace_text_with_format_and_images(self, excel_path: str, output_path: str, 
                                            text_data: List[Dict[str, Any]], translated_texts: List[str], 
                                            target_language: str = 'th') -> bool:
        """
        替换文本并保持格式和图片
        
        Args:
            excel_path: 输入Excel文件路径
            output_path: 输出Excel文件路径
            text_data: 原始文本数据
            translated_texts: 翻译后的文本列表
            target_language: 目标语言代码
            
        Returns:
            是否成功
        """
        try:
            workbook = load_workbook(excel_path, data_only=False)
            
            # 替换文本
            for item, translated_text in zip(text_data, translated_texts):
                sheet_name = item['sheet_name']
                row = item['row']
                column = item['column']
                format_info = item['format_info']
                
                # 获取工作表和单元格
                sheet = workbook[sheet_name]
                cell = sheet.cell(row=row, column=column)
                
                # 替换文本
                cell.value = translated_text
                
                # 应用格式
                self._apply_cell_format(cell, format_info, target_language)
                
                # 应用富文本格式（如果有）
                rich_text_info = item.get('rich_text_info')
                if rich_text_info and rich_text_info.get('has_rich_text'):
                    self._apply_rich_text_format(cell, item['text'], translated_text, rich_text_info, target_language)
                
                # 处理合并单元格同步（通用逻辑）
                merged_cell_info = self._check_merged_cell(cell)
                if merged_cell_info:
                    print(f"    📎 处理合并单元格: {merged_cell_info['range']}")
                    self._synchronize_merged_cell_formats(cell, item['text'], translated_text, format_info, rich_text_info, merged_cell_info)
                
                # 特别处理第78行 M-Q列（保持往后兼容性）
                if cell.row == 78 and cell.column >= 13 and cell.column <= 17:  # M=13, Q=17
                    print(f"    🔴 特别关注第78行 {cell.coordinate}")
                    print(f"      - 翻译前: '{item['text']}'")
                    print(f"      - 翻译后: '{translated_text}'")
                    print(f"      - 富文本信息: {rich_text_info}")
                    
                    # 如果没有检测到富文本但可能存在，尝试强制重新检查
                    if not rich_text_info:
                        print(f"      - 强制重新检查富文本格式...")
                        rich_text_info = self._extract_rich_text_format(cell)
                        if rich_text_info:
                            print(f"      - 重新检查发现富文本: {rich_text_info}")
                            self._apply_rich_text_format(cell, item['text'], translated_text, rich_text_info, target_language)
                            
                            # 如果发现了富文本且是合并单元格，重新同步
                            if merged_cell_info:
                                self._synchronize_merged_cell_formats(cell, item['text'], translated_text, format_info, rich_text_info, merged_cell_info)
                
                # 显示替换信息和格式应用情况
                color_applied = ""
                if format_info.get('font_color'):
                    if format_info.get('font_color_rgb'):
                        color_applied = f" [颜色:#{format_info['font_color_rgb']}]"
                    elif format_info.get('font_color_indexed'):
                        color_applied = f" [颜色:Indexed({format_info['font_color_indexed']})]"
                    elif format_info.get('font_color_theme'):
                        color_applied = f" [颜色:Theme({format_info['font_color_theme']})]"
                    else:
                        color_applied = " [颜色:已应用]"
                
                print(f"  替换文本: {sheet_name}!{cell.coordinate} = '{translated_text[:50]}...'{color_applied}")
            
            # 恢复图片信息
            print("🖼️ 恢复图片信息...")
            self.restore_images_info(workbook, self.image_data)
            
            # 保存文件
            workbook.save(output_path)
            workbook.close()
            
            print(f"✅ 成功替换 {len(text_data)} 个单元格的文本")
            return True
            
        except Exception as e:
            print(f"❌ 替换文本时发生错误: {e}")
            return False

    def _apply_cell_format(self, cell, format_info: Dict[str, Any], target_language: str = 'th') -> None:
        """
        应用单元格格式（增强泰文字体支持）
        
        Args:
            cell: openpyxl单元格对象
            format_info: 格式信息字典
            target_language: 目标语言代码
        """
        try:
            # 应用字体格式
            if format_info:
                font_kwargs = {}
                
                # 字体名称处理，优先支持泰文
                if target_language == 'th':
                    # 为泰文设置合适的字体
                    if format_info.get('font_name'):
                        # 保持原字体，但确保支持泰文
                        original_font = format_info['font_name']
                        # 常见的支持泰文的字体
                        thai_fonts = ['TH SarabunPSK', 'Tahoma', 'Arial Unicode MS', 'Microsoft Sans Serif']
                        
                        # 如果原字体在支持列表中，使用原字体；否则使用默认泰文字体
                        if any(thai_font.lower() in original_font.lower() for thai_font in thai_fonts):
                            font_kwargs['name'] = original_font
                        else:
                            font_kwargs['name'] = 'TH SarabunPSK'  # 默认泰文字体
                    else:
                        font_kwargs['name'] = 'TH SarabunPSK'
                elif format_info.get('font_name'):
                    font_kwargs['name'] = format_info['font_name']
                
                if format_info.get('font_size'):
                    # 应用字体大小调整
                    original_size = format_info['font_size']
                    adjusted_size = max(6, int(original_size * self.font_size_adjustment))
                    font_kwargs['size'] = adjusted_size
                
                if format_info.get('font_bold') is not None:
                    font_kwargs['bold'] = format_info['font_bold']
                
                if format_info.get('font_italic') is not None:
                    font_kwargs['italic'] = format_info['font_italic']
                
                if format_info.get('font_underline') is not None:
                    font_kwargs['underline'] = format_info['font_underline']
                
                if format_info.get('font_strike') is not None:
                    font_kwargs['strike'] = format_info['font_strike']
                
                # 改进颜色应用
                if format_info.get('font_color'):
                    try:
                        # 使用安全的颜色复制方法
                        safe_color = self._safe_copy_color(format_info['font_color'])
                        if safe_color:
                            font_kwargs['color'] = safe_color
                    except Exception as color_err:
                        print(f"⚠️ 设置字体颜色时出错: {color_err}")
                        # 如果有备用颜色信息，尝试使用
                        if format_info.get('font_color_rgb'):
                            try:
                                font_kwargs['color'] = Color(rgb=format_info['font_color_rgb'])
                            except Exception:
                                pass
                
                if font_kwargs:
                    try:
                        cell.font = Font(**font_kwargs)
                    except Exception as font_err:
                        print(f"⚠️ 设置字体格式失败: {font_err}")
                
                # 应用填充格式
                if format_info.get('fill_object'):
                    try:
                        # 直接使用原始填充对象
                        cell.fill = format_info['fill_object']
                    except Exception as fill_err:
                        print(f"⚠️ 使用原始填充对象失败: {fill_err}")
                        # 备用方案：使用分离的颜色和类型
                        if format_info.get('fill_color') and format_info.get('fill_type'):
                            try:
                                cell.fill = PatternFill(
                                    start_color=format_info['fill_color'],
                                    fill_type=format_info['fill_type']
                                )
                            except Exception as fill_err2:
                                print(f"⚠️ 备用填充方案也失败: {fill_err2}")
                
                # 应用对齐格式
                alignment_kwargs = {}
                if format_info.get('horizontal'):
                    alignment_kwargs['horizontal'] = format_info['horizontal']
                if format_info.get('vertical'):
                    alignment_kwargs['vertical'] = format_info['vertical']
                if format_info.get('wrap_text') is not None:
                    alignment_kwargs['wrap_text'] = format_info['wrap_text']
                if format_info.get('shrink_to_fit') is not None:
                    alignment_kwargs['shrink_to_fit'] = format_info['shrink_to_fit']
                
                if alignment_kwargs:
                    try:
                        cell.alignment = Alignment(**alignment_kwargs)
                    except Exception as align_err:
                        print(f"⚠️ 设置对齐格式失败: {align_err}")
                
                # 应用边框格式
                if format_info.get('border'):
                    try:
                        cell.border = format_info['border']
                    except Exception as border_err:
                        print(f"⚠️ 设置边框格式失败: {border_err}")
                
                # 应用数字格式
                if format_info.get('number_format'):
                    try:
                        cell.number_format = format_info['number_format']
                    except Exception as num_err:
                        print(f"⚠️ 设置数字格式失败: {num_err}")
                        
                # 调试信息：显示应用的格式（只在有颜色时显示）
                if format_info.get('font_color') and (format_info.get('font_color_rgb') or format_info.get('font_color_indexed') or format_info.get('font_color_theme')):
                    color_info = ""
                    if format_info.get('font_color_rgb'):
                        color_info = f"RGB: #{format_info['font_color_rgb']}"
                    elif format_info.get('font_color_indexed'):
                        color_info = f"Indexed: {format_info['font_color_indexed']}"
                    elif format_info.get('font_color_theme'):
                        color_info = f"Theme: {format_info['font_color_theme']}"
                        if format_info.get('font_color_tint'):
                            color_info += f" Tint: {format_info['font_color_tint']}"
                    print(f"    🎨 应用字体颜色: {color_info}")
                    
        except Exception as e:
            print(f"⚠️ 应用格式时出错: {e}")

    def smart_adjust_column_width(self, excel_path: str, output_path: Optional[str] = None) -> None:
        """
        智能调整列宽以适应内容，同时考虑图片位置
        
        Args:
            excel_path: Excel文件路径
            output_path: 输出文件路径，如果为None则覆盖原文件
        """
        try:
            if output_path is None:
                output_path = excel_path
            
            workbook = load_workbook(excel_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 获取该工作表的图片信息
                sheet_images = self.image_data.get(sheet_name, [])
                occupied_columns = set()
                
                # 标记被图片占用的列
                for img_info in sheet_images:
                    anchor_info = img_info.get('anchor_info', {})
                    if anchor_info.get('type') == 'two_cell':
                        from_col = anchor_info.get('from_col', 0)
                        to_col = anchor_info.get('to_col', 0)
                        for col in range(from_col, to_col + 1):
                            occupied_columns.add(col)
                
                # 遍历所有列
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    column_index = column[0].column
                    
                    # 如果该列被图片占用，使用更保守的宽度调整
                    is_occupied = column_index in occupied_columns
                    
                    for cell in column:
                        try:
                            if cell.value:
                                # 计算单元格内容长度
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except Exception:
                            pass
                    
                    # 设置列宽，考虑图片占用情况
                    if is_occupied:
                        # 图片占用的列使用更保守的宽度
                        adjusted_width = min(max_length + 1, 30)
                    else:
                        # 正常列使用标准宽度
                        adjusted_width = min(max_length + 2, 50)
                    
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(output_path)
            workbook.close()
            print(f"✅ 已智能调整列宽")
            
        except Exception as e:
            print(f"❌ 调整列宽时发生错误: {e}")

    def analyze_excel_structure(self, excel_path: str) -> Dict[str, Any]:
        """
        分析Excel文件结构，包括图片和文本分布
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            分析结果字典
        """
        try:
            workbook = load_workbook(excel_path, data_only=False)
            analysis = {
                'sheets': {},
                'total_images': 0,
                'total_text_cells': 0
            }
            
            # 提取图片信息
            image_data = self.extract_images_info(workbook)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info = {
                    'name': sheet_name,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'images': len(image_data.get(sheet_name, [])),
                    'text_cells': 0,
                    'chinese_cells': 0
                }
                
                # 统计文本单元格
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.strip():
                            sheet_info['text_cells'] += 1
                            if re.search(r'[\u4e00-\u9fff]', cell.value):
                                sheet_info['chinese_cells'] += 1
                
                analysis['sheets'][sheet_name] = sheet_info
                analysis['total_images'] += sheet_info['images']
                analysis['total_text_cells'] += sheet_info['text_cells']
            
            workbook.close()
            
            # 打印分析结果
            print("📊 Excel文件结构分析")
            print("=" * 50)
            print(f"总工作表数: {len(analysis['sheets'])}")
            print(f"总图片数: {analysis['total_images']}")
            print(f"总文本单元格数: {analysis['total_text_cells']}")
            print("\n各工作表详情:")
            for sheet_name, info in analysis['sheets'].items():
                print(f"  📄 {sheet_name}:")
                print(f"    - 尺寸: {info['max_row']}行 x {info['max_column']}列")
                print(f"    - 图片: {info['images']}个")
                print(f"    - 文本单元格: {info['text_cells']}个")
                print(f"    - 中文单元格: {info['chinese_cells']}个")
            
            return analysis
            
        except Exception as e:
            print(f"❌ 分析Excel结构时发生错误: {e}")
            return {}


def main() -> None:
    """主函数 - 演示如何使用ExcelTranslatorV2"""
    # 创建翻译器实例
    translator = ExcelTranslatorV2(font_size_adjustment=0.8)
    
    # 设置文件路径
    input_file = "test.xlsx"  # 输入Excel文件
    output_file = "output_translated_thai.xlsx"  # 输出Excel文件
    
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"❌ 输入文件不存在: {input_file}")
        return
    
    # 分析Excel文件结构
    print("🔍 分析Excel文件结构...")
    analysis = translator.analyze_excel_structure(input_file)
    
    if analysis.get('total_images', 0) > 0:
        print(f"✅ 检测到 {analysis['total_images']} 个图片，将使用V2版本保护图片")
    
    # 执行翻译
    print("🚀 开始翻译Excel文件...")
    success = translator.replace_text_in_excel(
        excel_path=input_file,
        output_path=output_file,
        target_language='th'  # 翻译为泰文
    )
    
    if success:
        # 智能调整列宽
        print("📐 智能调整列宽...")
        translator.smart_adjust_column_width(output_file)
        
        print("🎉 翻译完成！")
        print(f"输入文件: {input_file}")
        print(f"输出文件: {output_file}")
        print("\n✨ V2版本新特性:")
        print("- 🖼️ 完整保护Excel中的图片不变形")
        print("- 📏 智能调整列宽避免影响图片布局")
        print("- 🎯 改进的锚点位置保持")
        print("- 🔍 详细的文件结构分析")
    else:
        print("❌ 翻译失败")


if __name__ == "__main__":
    print("Excel翻译工具 V2 - 图片保护版")
    print("=" * 50)
    print("V2版本新特性:")
    print("1. ✅ 支持翻译Excel中的所有文本")
    print("2. ✅ 保持原有格式（字体、颜色、对齐等）")
    print("3. ✅ 支持多工作表")
    print("4. ✅ 自动调整字体大小")
    print("5. 🆕 完整保护图片不变形")
    print("6. 🆕 智能列宽调整")
    print("7. 🆕 图片锚点位置保持")
    print("8. 🆕 Excel结构分析")
    print("9. ✅ 跳过公式单元格")
    print("10. ✅ 批量翻译提高效率")
    print("11. 🆕 智能文本过滤和去重")
    print("12. 🆕 API调用优化和统计")
    print("13. 🆕 增强字体颜色保留")
    print("14. 🆕 单元格内混合颜色富文本支持")
    print("15. 🆕 合并单元格富文本增强支持")
    print("=" * 50)
    main()