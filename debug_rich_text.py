#!/usr/bin/env python3
"""
Debug rich text extraction specifically
"""

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
import logging

logging.basicConfig(level=logging.DEBUG)

def debug_rich_text_cells(file_path: str):
    """Debug rich text extraction"""
    print(f"=== Debugging Rich Text in: {file_path} ===\n")
    
    workbook = load_workbook(file_path, data_only=False)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Sheet: {sheet_name}")
        
        # Check specific cells that should have rich text
        rich_text_cells = ['A3', 'B2', 'N78']  # Based on our test file
        
        for cell_addr in rich_text_cells:
            if cell_addr in [cell.coordinate for row in sheet.iter_rows() for cell in row]:
                cell = sheet[cell_addr]
                print(f"\nCell {cell_addr}:")
                print(f"  Value: {cell.value}")
                print(f"  Value type: {type(cell.value)}")
                print(f"  _value type: {type(cell._value) if hasattr(cell, '_value') else 'None'}")
                
                # Check if it's rich text
                is_rich_text_value = isinstance(cell.value, CellRichText)
                is_rich_text_internal = hasattr(cell, '_value') and isinstance(cell._value, CellRichText)
                
                print(f"  Is rich text (value): {is_rich_text_value}")
                print(f"  Is rich text (_value): {is_rich_text_internal}")
                
                if is_rich_text_value:
                    print(f"  Rich text segments in value: {len(list(cell.value))}")
                    for i, segment in enumerate(cell.value):
                        print(f"    Segment {i}: {type(segment)} = {segment}")
                        if hasattr(segment, 'font'):
                            print(f"      Font: {segment.font}")
                
                if is_rich_text_internal:
                    print(f"  Rich text segments in _value: {len(list(cell._value))}")
                    for i, segment in enumerate(cell._value):
                        print(f"    Segment {i}: {type(segment)} = {segment}")
                        if hasattr(segment, 'font'):
                            print(f"      Font: {segment.font}")
                
                # Check for any other rich text attributes
                rich_attrs = [attr for attr in dir(cell) if 'rich' in attr.lower()]
                if rich_attrs:
                    print(f"  Rich text related attributes: {rich_attrs}")
                    for attr in rich_attrs:
                        try:
                            value = getattr(cell, attr)
                            if value:
                                print(f"    {attr}: {type(value)} = {value}")
                        except Exception:
                            pass
            else:
                print(f"\nCell {cell_addr}: Not found")
    
    workbook.close()

if __name__ == "__main__":
    debug_rich_text_cells("/root/projects/github/Offitrans/complex_test.xlsx")