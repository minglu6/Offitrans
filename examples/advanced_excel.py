#!/usr/bin/env python3
"""
Advanced Excel Translation Examples for Offitrans

This example demonstrates advanced Excel translation features including:
- Rich text formatting preservation
- Image protection
- Custom configuration
- Error handling
- Statistics and monitoring
"""

import os
import logging
from pathlib import Path

from offitrans import ExcelProcessor, GoogleTranslator
from offitrans.core.config import Config
from offitrans.exceptions.errors import ExcelProcessorError

# Set up detailed logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def advanced_excel_with_images():
    """
    Example of Excel translation with image protection
    """
    print("=" * 60)
    print("Advanced Excel Translation with Image Protection")
    print("=" * 60)
    
    # Create advanced configuration
    config = Config()
    
    # Translator settings
    config.translator.max_workers = 3
    config.translator.timeout = 30
    config.translator.retry_count = 2
    
    # Processor settings
    config.processor.preserve_formatting = True
    config.processor.image_protection = True  # Enable image protection
    config.processor.font_size_adjustment = 0.8
    config.processor.smart_column_width = True
    
    # Cache settings
    config.cache.enabled = True
    config.cache.cache_file = "advanced_translation_cache.json"
    config.cache.auto_save_interval = 5
    
    # Create custom translator
    translator = GoogleTranslator(
        source_lang="zh",
        target_lang="en", 
        use_free_api=True,
        **config.get_translator_kwargs()
    )
    
    # Create Excel processor
    processor = ExcelProcessor(
        translator=translator,
        config=config
    )
    
    # File paths
    input_file = "examples/sample_files/excel_with_images.xlsx"
    output_file = "examples/sample_files/excel_with_images_translated.xlsx"
    
    print(f"ğŸ“Š Processing Excel file with image protection...")
    print(f"   Input: {input_file}")
    print(f"   Output: {output_file}")
    
    if not os.path.exists(input_file):
        print(f"âš ï¸  Sample file not found: {input_file}")
        print("   Create an Excel file with Chinese text and images to test this feature")
        create_sample_excel_with_formatting(input_file)
        return
    
    try:
        # Process the file
        success = processor.process_file(input_file, output_file, "en")
        
        if success:
            print("âœ… Translation completed with image protection!")
            
            # Show detailed statistics
            stats = processor.get_stats()
            print(f"\nğŸ“ˆ Processing Statistics:")
            print(f"   Files processed: {stats['total_files_processed']}")
            print(f"   Successful files: {stats['successful_files']}")
            print(f"   Failed files: {stats['failed_files']}")
            print(f"   Texts translated: {stats['total_texts_translated']}")
            print(f"   Characters translated: {stats['total_chars_translated']}")
            
            # Show translator statistics
            translator_stats = translator.get_stats()
            print(f"\nğŸ”¤ Translator Statistics:")
            print(f"   Total translations: {translator_stats['total_translations']}")
            print(f"   Successful translations: {translator_stats['successful_translations']}")
            print(f"   Failed translations: {translator_stats['failed_translations']}")
            
        else:
            print("âŒ Translation failed")
            
    except ExcelProcessorError as e:
        print(f"âŒ Excel processing error: {e}")
        if e.file_path:
            print(f"   File: {e.file_path}")
        if e.details:
            print(f"   Details: {e.details}")
    except Exception as e:
        print(f"âŒ Unexpected error: {e}")


def batch_excel_translation():
    """
    Example of batch Excel file translation
    """
    print("\n" + "=" * 60)
    print("Batch Excel Translation Example")
    print("=" * 60)
    
    # List of Excel files to translate
    excel_files = [
        "examples/sample_files/file1.xlsx",
        "examples/sample_files/file2.xlsx", 
        "examples/sample_files/file3.xlsx"
    ]
    
    # Create processor
    config = Config()
    config.translator.max_workers = 2  # Conservative for batch processing
    
    processor = ExcelProcessor(config=config)
    
    # Process each file
    successful_files = []
    failed_files = []
    
    print(f"ğŸ“Š Processing {len(excel_files)} Excel files...")
    
    for i, input_file in enumerate(excel_files, 1):
        output_file = input_file.replace('.xlsx', '_translated.xlsx')
        
        print(f"\nğŸ“„ Processing file {i}/{len(excel_files)}: {input_file}")
        
        if not os.path.exists(input_file):
            print(f"   âš ï¸  File not found, creating sample...")
            create_sample_excel_with_formatting(input_file)
        
        try:
            success = processor.process_file(input_file, output_file, "en")
            
            if success:
                print(f"   âœ… Success: {output_file}")
                successful_files.append(output_file)
            else:
                print(f"   âŒ Failed")
                failed_files.append(input_file)
                
        except Exception as e:
            print(f"   âŒ Error: {e}")
            failed_files.append(input_file)
    
    # Summary
    print(f"\nğŸ“Š Batch Processing Summary:")
    print(f"   Total files: {len(excel_files)}")
    print(f"   Successful: {len(successful_files)}")
    print(f"   Failed: {len(failed_files)}")
    
    if successful_files:
        print(f"\nâœ… Successfully translated files:")
        for file in successful_files:
            print(f"   - {file}")
    
    if failed_files:
        print(f"\nâŒ Failed files:")
        for file in failed_files:
            print(f"   - {file}")


def excel_translation_with_custom_formatting():
    """
    Example of Excel translation with custom formatting options
    """
    print("\n" + "=" * 60)
    print("Excel Translation with Custom Formatting")
    print("=" * 60)
    
    # Test different font size adjustments
    font_adjustments = [0.6, 0.8, 1.0, 1.2]
    
    base_input = "examples/sample_files/formatting_test.xlsx"
    
    if not os.path.exists(base_input):
        print(f"Creating sample file with formatting: {base_input}")
        create_sample_excel_with_formatting(base_input)
    
    for adjustment in font_adjustments:
        print(f"\nğŸ”§ Testing font size adjustment: {adjustment}")
        
        # Create config with specific font adjustment
        config = Config()
        config.processor.font_size_adjustment = adjustment
        
        processor = ExcelProcessor(config=config)
        
        output_file = f"examples/sample_files/formatting_test_adj_{adjustment}.xlsx"
        
        try:
            success = processor.process_file(base_input, output_file, "en")
            
            if success:
                print(f"   âœ… Created: {output_file}")
            else:
                print(f"   âŒ Failed to create: {output_file}")
                
        except Exception as e:
            print(f"   âŒ Error: {e}")


def excel_translation_different_languages():
    """
    Example of translating Excel to different languages
    """
    print("\n" + "=" * 60)
    print("Excel Translation to Multiple Languages")
    print("=" * 60)
    
    input_file = "examples/sample_files/multilang_source.xlsx"
    
    if not os.path.exists(input_file):
        print(f"Creating sample multilingual source file: {input_file}")
        create_sample_excel_with_formatting(input_file)
    
    # Different target languages
    languages = {
        "en": "English",
        "th": "Thai",
        "ja": "Japanese", 
        "fr": "French"
    }
    
    config = Config()
    config.translator.max_workers = 1  # Conservative for multiple languages
    
    for lang_code, lang_name in languages.items():
        print(f"\nğŸŒ Translating to {lang_name} ({lang_code})...")
        
        # Create language-specific translator
        translator = GoogleTranslator(
            source_lang="zh",
            target_lang=lang_code,
            use_free_api=True,
            max_workers=1
        )
        
        processor = ExcelProcessor(translator=translator, config=config)
        
        output_file = f"examples/sample_files/multilang_output_{lang_code}.xlsx"
        
        try:
            success = processor.process_file(input_file, output_file, lang_code)
            
            if success:
                print(f"   âœ… {lang_name} translation completed: {output_file}")
            else:
                print(f"   âŒ {lang_name} translation failed")
                
        except Exception as e:
            print(f"   âŒ {lang_name} translation error: {e}")


def create_sample_excel_with_formatting(file_path: str):
    """
    Create a sample Excel file with various formatting for testing
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.styles.colors import Color
        
        # Create parent directory if it doesn't exist
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "æµ‹è¯•å·¥ä½œè¡¨"
        
        # Sample data with different formatting
        data = [
            ("æ ‡é¢˜", "è¿™æ˜¯ä¸€ä¸ªæ ‡é¢˜è¡Œ"),
            ("æ™®é€šæ–‡æœ¬", "è¿™æ˜¯æ™®é€šçš„ä¸­æ–‡æ–‡æœ¬"),
            ("é‡è¦ä¿¡æ¯", "è¿™æ˜¯åŠ ç²—çš„é‡è¦ä¿¡æ¯"),
            ("æ•°å­—", "123"),  # Should not be translated
            ("é‚®ç®±", "test@example.com"),  # Should not be translated
            ("æ··åˆå†…å®¹", "åŒ…å«ä¸­æ–‡å’ŒEnglishçš„æ··åˆæ–‡æœ¬"),
            ("é•¿æ–‡æœ¬", "è¿™æ˜¯ä¸€ä¸ªå¾ˆé•¿çš„æ–‡æœ¬å†…å®¹ï¼Œç”¨äºæµ‹è¯•è‡ªåŠ¨æ¢è¡Œå’Œåˆ—å®½è°ƒæ•´åŠŸèƒ½ã€‚å®ƒåŒ…å«äº†å¤šä¸ªå¥å­å’Œå„ç§ä¸­æ–‡å­—ç¬¦ã€‚")
        ]
        
        for i, (category, content) in enumerate(data, 1):
            # Category column with bold
            cell_a = ws[f'A{i}']
            cell_a.value = category
            cell_a.font = Font(bold=True, size=12)
            cell_a.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
            
            # Content column
            cell_b = ws[f'B{i}']
            cell_b.value = content
            
            # Different formatting for different rows
            if i == 1:  # Title row
                cell_b.font = Font(bold=True, size=14, color="FF0000FF")
            elif i == 3:  # Important info
                cell_b.font = Font(bold=True, italic=True, color="FFFF0000")
            elif i == 7:  # Long text
                cell_b.alignment = Alignment(wrap_text=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        
        wb.save(file_path)
        print(f"âœ… Created sample Excel file: {file_path}")
        
    except ImportError:
        print("âŒ openpyxl not available, cannot create sample Excel file")
    except Exception as e:
        print(f"âŒ Error creating sample Excel file: {e}")


def main():
    """
    Main function to run all advanced examples
    """
    print("ğŸš€ Offitrans Advanced Excel Translation Examples")
    print("This example demonstrates advanced Excel translation features")
    
    # Ensure sample files directory exists
    os.makedirs("examples/sample_files", exist_ok=True)
    
    # Run all examples
    advanced_excel_with_images()
    batch_excel_translation()
    excel_translation_with_custom_formatting()
    excel_translation_different_languages()
    
    print("\n" + "=" * 60)
    print("âœ¨ Advanced examples completed!")
    print("=" * 60)
    print("ğŸ’¡ Advanced Tips:")
    print("   1. Use image_protection=True for Excel files with images")
    print("   2. Adjust font_size_adjustment for better text fit")
    print("   3. Enable caching for better performance with repeated translations")
    print("   4. Use custom configurations for different use cases")
    print("   5. Monitor statistics to track translation performance")
    print("\nğŸ“š Check the generated files in examples/sample_files/")


if __name__ == "__main__":
    main()